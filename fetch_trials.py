#!/usr/bin/env python3
"""
ClinicalTrials.gov Lead Generation Dashboard Generator
=======================================================
TWO MODES:
  1. API mode  — fetches live data directly from clinicaltrials.gov (run locally)
  2. CSV mode  — reads a downloaded CSV/XLSX export from clinicaltrials.gov

Usage:
    python fetch_trials.py                          # live API fetch, last 365 days (default)
    python fetch_trials.py --days 180               # last 6 months only
    python fetch_trials.py --days 30                # last 30 days (quick pulse check)
    python fetch_trials.py --test                   # 1-page API test (~1000 records)
    python fetch_trials.py --max-pages 5            # cap at 5 API pages
    python fetch_trials.py --csv trials.csv         # process a downloaded CSV
    python fetch_trials.py --csv trials.xlsx        # process a downloaded XLSX
    python fetch_trials.py --use-llm                # force LLM classification (needs ANTHROPIC_API_KEY)
    python fetch_trials.py --no-llm                 # force keyword-only classification
    python fetch_trials.py --cache-path cache.json  # custom classification cache
"""

import sys, json, time, argparse, re, csv, io, os
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    import requests
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests",
                           "--break-system-packages", "-q"])
    import requests

# ─── OUTPUT ────────────────────────────────────────────────────────────────────
OUTPUT_HTML = Path("clinical_trials_dashboard.html")

# ─── API CONFIG ────────────────────────────────────────────────────────────────
API_BASE        = "https://clinicaltrials.gov/api/v2/studies"
PAGE_SIZE       = 1000
REQUEST_DELAY   = 0.4   # seconds between pages – polite crawling
REQUEST_TIMEOUT = 30

TARGET_STATUSES = [
    "RECRUITING",
    "ACTIVE_NOT_RECRUITING",
    "NOT_YET_RECRUITING",
    "ENROLLING_BY_INVITATION",
]

# ─── THERAPY AREA KEYWORD MAP ──────────────────────────────────────────────────
# Matched against: conditions + brief title + summary + interventions (multi-tag)
THERAPY_AREAS = {
    "Oncology": [
        # Core tumor types
        "cancer", "carcinoma", "tumor", "tumour", "lymphoma", "leukemia",
        "leukaemia", "melanoma", "sarcoma", "myeloma", "glioma", "glioblastoma",
        "blastoma", "adenocarcinoma", "malignant neoplasm", "metastatic",
        "nsclc", "sclc", "hepatocellular", "cholangiocarcinoma", "mesothelioma",
        "neuroblastoma", "osteosarcoma", "rhabdomyosarcoma", "wilms tumor",
        "retinoblastoma", "thymoma", "kaposi sarcoma", "merkel cell",
        "pheochromocytoma", "paraganglioma",
        # Extended cancer subtypes
        "renal cell carcinoma", "renal cell cancer", "urothelial carcinoma",
        "bladder cancer", "prostate cancer", "colorectal cancer", "colon cancer",
        "rectal cancer", "pancreatic cancer", "gastric cancer", "stomach cancer",
        "esophageal cancer", "esophagogastric", "head and neck cancer",
        "squamous cell", "thyroid cancer", "medullary thyroid", "anaplastic thyroid",
        "diffuse large b-cell", "dlbcl", "follicular lymphoma", "mantle cell lymphoma",
        "burkitt lymphoma", "hodgkin lymphoma", "peripheral t-cell lymphoma",
        "chronic lymphocytic leukemia", "cll", "chronic myeloid leukemia", "cml",
        "acute myeloid leukemia", "aml", "acute lymphoblastic leukemia", "all",
        "myelofibrosis", "myelodysplastic", "myeloproliferative",
        "mycosis fungoides", "sezary syndrome",
        # Biomarkers & actionable mutations
        "kras g12", "egfr mutation", "egfr exon", "her2-positive", "her2-low",
        "her2 amplification", "alk-positive", "alk positive", "ros1 fusion",
        "braf v600", "braf mutation", "pd-l1", "brca mutation", "brca1", "brca2",
        "msi-h", "microsatellite instability", "tmb-h", "tumor mutational burden",
        "met amplification", "met exon 14", "fgfr alteration", "fgfr2", "fgfr3",
        "ntrk fusion", "ret fusion", "erbb2", "pik3ca mutation", "kras mutation",
        # Treatment classes
        "checkpoint inhibitor", "car-t cell", "chimeric antigen receptor",
        "bispecific antibody", "antibody-drug conjugate", "adc therapy",
        "parp inhibitor", "cdk4/6 inhibitor", "bcl-2 inhibitor", "btk inhibitor",
        "radioligand therapy", "lutetium-177", "radium-223",
        # High-frequency drug keywords (appear in titles/summaries)
        "pembrolizumab", "nivolumab", "atezolizumab", "durvalumab", "cemiplimab",
        "ipilimumab", "tremelimumab", "relatlimab", "dostarlimab",
        "trastuzumab deruxtecan", "sacituzumab govitecan", "enfortumab vedotin",
        "mirvetuximab", "patritumab deruxtecan", "datopotamab deruxtecan",
        "rituximab", "obinutuzumab", "daratumumab", "isatuximab",
        "teclistamab", "talquetamab", "elranatamab", "linvoseltamab",
        "ciltacabtagene", "idecabtagene", "tisagenlecleucel", "axicabtagene",
        "lisocabtagene", "brexucabtagene",
        "venetoclax", "ibrutinib", "acalabrutinib", "zanubrutinib", "pirtobrutinib",
        "palbociclib", "ribociclib", "abemaciclib",
        "olaparib", "niraparib", "rucaparib", "talazoparib", "fuzuloparib",
        "adagrasib", "sotorasib", "divarasib",
        "osimertinib", "lazertinib", "erlotinib", "gefitinib", "afatinib",
        "alectinib", "brigatinib", "lorlatinib", "crizotinib",
        "cabozantinib", "lenvatinib", "sorafenib", "regorafenib", "sunitinib",
        "imatinib", "dasatinib", "nilotinib", "ponatinib", "asciminib",
        "midostaurin", "gilteritinib", "quizartinib", "enasidenib", "ivosidenib",
        "azacitidine", "decitabine", "bortezomib", "carfilzomib", "ixazomib",
        "alpelisib", "copanlisib", "idelalisib", "umbralisib",
        "tarlatamab", "amivantamab", "telisotuzumab",
        "mogamulizumab", "brentuximab vedotin", "polatuzumab vedotin",
        "inotuzumab ozogamicin", "gemtuzumab ozogamicin",
    ],
    "Neurology": [
        # Core conditions
        "alzheimer", "parkinson", "multiple sclerosis", "epilepsy",
        "seizure disorder", "stroke", "cerebrovascular", "dementia",
        "peripheral neuropathy", "amyotrophic lateral sclerosis",
        "huntington", "spinal muscular atrophy", "migraine",
        "narcolepsy", "prion disease", "frontotemporal dementia",
        "spinocerebellar", "charcot-marie-tooth", "guillain-barre",
        "myasthenia gravis", "neuromyelitis optica", "transient ischemic attack",
        # Extended conditions
        "lewy body dementia", "vascular dementia", "mixed dementia",
        "traumatic brain injury", "tbi", "post-stroke", "ischemic stroke",
        "hemorrhagic stroke", "subarachnoid hemorrhage",
        "relapsing remitting ms", "progressive ms", "primary progressive ms",
        "secondary progressive ms", "clinically isolated syndrome",
        "dravet syndrome", "lennox-gastaut", "tuberous sclerosis",
        "neurofibromatosis", "spinal cord", "friedreich ataxia",
        "progressive supranuclear palsy", "psp", "corticobasal syndrome",
        "primary lateral sclerosis", "essential tremor", "restless legs syndrome",
        "normal pressure hydrocephalus", "idiopathic intracranial hypertension",
        "chronic inflammatory demyelinating polyneuropathy", "cidp",
        "multifocal motor neuropathy", "small fiber neuropathy",
        "hereditary spastic paraplegia", "rett syndrome neurology",
        "angelman syndrome", "focal epilepsy", "generalised epilepsy",
        "status epilepticus", "infantile spasms", "west syndrome",
        # Biomarkers & mechanisms
        "amyloid beta", "tau protein", "alpha-synuclein", "tdp-43",
        "sod1 mutation", "fus mutation", "c9orf72",
        "cgrp", "cgrp receptor", "anti-cd20 neurology",
        # Drug keywords
        "lecanemab", "donanemab", "aducanumab", "gantenerumab",
        "ocrelizumab", "ofatumumab", "natalizumab", "siponimod",
        "ozanimod", "ponesimod", "cladribine", "alemtuzumab",
        "dimethyl fumarate", "teriflunomide", "fingolimod",
        "erenumab", "fremanezumab", "galcanezumab", "eptinezumab",
        "ubrogepant", "rimegepant", "atogepant", "lasmiditan",
        "safinamide", "opicapone", "prasinezumab", "nusinersen",
        "onasemnogene abeparvovec", "risdiplam",
    ],
    "Cardiovascular": [
        # Core conditions
        "heart failure", "cardiac", "cardiovascular", "hypertension",
        "atrial fibrillation", "coronary artery disease", "myocardial infarction",
        "arrhythmia", "cardiomyopathy", "peripheral artery disease",
        "aortic stenosis", "pulmonary arterial hypertension",
        "ventricular dysfunction", "angina", "aortic aneurysm",
        "carotid artery", "atherosclerosis", "pulmonary embolism",
        # Extended conditions
        "atrial flutter", "ventricular tachycardia", "ventricular fibrillation",
        "sudden cardiac death", "reduced ejection fraction", "hfref",
        "preserved ejection fraction", "hfpef", "hfmref",
        "hypertrophic cardiomyopathy", "dilated cardiomyopathy",
        "cardiac amyloidosis", "transthyretin cardiac",
        "peripheral vascular disease", "critical limb ischemia",
        "deep vein thrombosis", "dvt", "thromboembolism",
        "hyperlipidemia", "hypercholesterolemia", "dyslipidemia",
        "familial hypercholesterolemia", "hypertriglyceridemia",
        "acute coronary syndrome", "stemi", "nstemi",
        "cardiac resynchronization", "implantable cardioverter",
        "transcatheter aortic", "tavi", "tavr", "mitral regurgitation",
        "ldl cholesterol", "triglycerides",
        # Drug keywords
        "sacubitril", "valsartan entresto", "ivabradine",
        "rivaroxaban", "apixaban", "dabigatran", "edoxaban",
        "ticagrelor", "clopidogrel", "prasugrel",
        "evolocumab", "alirocumab", "inclisiran", "bempedoic acid",
        "icosapentaenoic acid", "ezetimibe",
        "empagliflozin heart", "dapagliflozin heart",
        "tafamidis", "acoramidis", "mavacamten", "omecamtiv mecarbil",
        "vericiguat", "abelacimab", "asundexian", "milvexian",
    ],
    "Immunology & Inflammation": [
        # Core conditions
        "systemic lupus", "rheumatoid arthritis", "psoriatic arthritis",
        "crohn", "ulcerative colitis", "inflammatory bowel disease",
        "sjogren", "vasculitis", "ankylosing spondylitis",
        "giant cell arteritis", "polymyalgia rheumatica", "myositis",
        "dermatomyositis", "systemic sclerosis", "scleroderma",
        "graft-versus-host", "gvhd", "behcet", "antiphospholipid",
        "primary immunodeficiency", "mastocytosis",
        # Extended conditions
        "axial spondyloarthritis", "nr-axspa", "non-radiographic",
        "peripheral spondyloarthritis", "reactive arthritis", "enthesitis",
        "anca-associated vasculitis", "granulomatosis with polyangiitis",
        "eosinophilic granulomatosis", "polyarteritis nodosa",
        "iga vasculitis", "anti-gbm disease", "glomerulonephritis autoimmune",
        "adult-onset still", "systemic juvenile idiopathic arthritis",
        "mixed connective tissue", "undifferentiated connective tissue",
        "overlap syndrome autoimmune", "antinuclear antibody",
        "complement deficiency", "hereditary periodic fever", "nlrp3",
        "autoinflammatory", "familial mediterranean fever",
        "lupus nephritis", "cutaneous lupus", "discoid lupus",
        "type i interferonopathy", "aicardi-goutieres",
        # Drug keywords
        "adalimumab", "etanercept", "infliximab", "golimumab", "certolizumab",
        "ustekinumab", "secukinumab", "ixekizumab", "bimekizumab",
        "guselkumab", "risankizumab", "tildrakizumab",
        "abatacept", "tocilizumab", "sarilumab", "olokizumab",
        "baricitinib", "upadacitinib", "filgotinib", "tofacitinib",
        "anifrolumab", "belimumab", "voclosporin", "avacopan",
        "deucravacitinib", "spesolimab", "imsidolimab",
        "vedolizumab", "ozanimod ibd", "etrasimod", "mirikizumab ibd",
    ],
    "Infectious Disease": [
        # Core conditions
        "hiv infection", "hiv/aids", "tuberculosis", "malaria",
        "hepatitis b", "hepatitis c", "covid-19", "sars-cov",
        "influenza", "bacterial infection", "viral infection",
        "fungal infection", "sepsis", "pneumonia", "respiratory syncytial",
        "dengue fever", "ebola", "human papillomavirus", "cytomegalovirus",
        "clostridium difficile", "staphylococcus aureus", "lyme disease",
        "chagas", "antimicrobial resistance",
        # Extended conditions
        "hiv-1", "hiv pre-exposure prophylaxis", "prep hiv",
        "antiretroviral", "highly active antiretroviral", "haart",
        "mpox", "monkeypox", "rsv infection", "respiratory syncytial virus",
        "meningitis", "encephalitis", "meningococcal disease",
        "hospital-acquired infection", "nosocomial infection",
        "surgical site infection", "bloodstream infection",
        "invasive candidiasis", "candida infection", "aspergillosis",
        "cryptococcal meningitis", "mucormycosis",
        "antibiotic resistance", "carbapenem-resistant", "mdr-tb",
        "extensively drug-resistant", "xdr",
        "dengue", "zika virus", "west nile virus",
        "leishmaniasis", "schistosomiasis", "hookworm", "trypanosomiasis",
        "onchocerciasis", "lymphatic filariasis",
        "herpes simplex", "varicella zoster", "epstein-barr",
        "clostridioides difficile", "cdiff", "c. diff",
        # Drug keywords
        "lenacapavir", "cabotegravir", "islatravir",
        "dolutegravir", "biktarvy", "elvitegravir",
        "nirmatrelvir", "ritonavir covid", "remdesivir", "molnupiravir",
        "glecaprevir", "pibrentasvir", "sofosbuvir", "velpatasvir",
        "ceftazidime-avibactam", "cefiderocol", "imipenem-relebactam",
        "meropenem-vaborbactam", "aztreonam-avibactam",
        "isavuconazole", "voriconazole", "posaconazole", "ibrexafungerp",
        "olorofim", "fosmanogepix",
    ],
    "Respiratory": [
        # Core conditions
        "asthma", "chronic obstructive pulmonary", "idiopathic pulmonary fibrosis",
        "lung fibrosis", "bronchiectasis", "cystic fibrosis",
        "interstitial lung disease", "allergic rhinitis",
        "chronic bronchitis", "emphysema", "obstructive sleep apnea",
        "eosinophilic asthma", "hypersensitivity pneumonitis", "sarcoidosis",
        # Extended conditions
        "copd", "severe asthma", "uncontrolled asthma", "refractory asthma",
        "exercise-induced bronchoconstriction", "vocal cord dysfunction",
        "pulmonary hypertension", "chronic thromboembolic pulmonary hypertension",
        "cteph", "group 1 pah", "pulmonary vascular",
        "non-cf bronchiectasis", "allergic bronchopulmonary aspergillosis",
        "acute exacerbation of copd", "aecopd",
        "lung transplant", "post-transplant bronchiolitis obliterans",
        "interstitial pneumonia", "usual interstitial pneumonia", "uip",
        "desquamative interstitial pneumonia", "cryptogenic organizing pneumonia",
        "acute respiratory distress", "ards",
        "alpha-1 antitrypsin",
        # Drug keywords
        "tezepelumab", "mepolizumab", "benralizumab", "omalizumab",
        "dupilumab asthma", "reslizumab",
        "nintedanib", "pirfenidone", "pamrevlumab",
        "elexacaftor", "tezacaftor", "ivacaftor", "lumacaftor",
        "sotatercept pulmonary", "selexipag", "ambrisentan",
        "bosentan", "macitentan", "riociguat", "sildenafil pulmonary",
        "tiotropium", "umeclidinium", "indacaterol", "glycopyrrolate inhaler",
        "salmeterol", "formoterol", "budesonide inhaled", "fluticasone inhaled",
    ],
    "Metabolic & Endocrine": [
        # Core conditions
        "type 2 diabetes", "type 1 diabetes", "obesity", "overweight",
        "hypothyroidism", "hyperthyroidism", "graves disease", "cushing syndrome",
        "metabolic syndrome", "insulin resistance", "hypercholesterolemia",
        "nonalcoholic steatohepatitis", "nash", "nafld",
        "non-alcoholic fatty liver", "growth hormone deficiency",
        "acromegaly", "hyperparathyroidism", "adrenal insufficiency",
        # Extended conditions
        "diabetes mellitus", "prediabetes", "impaired fasting glucose",
        "impaired glucose tolerance", "hyperglycemia", "hypoglycemia",
        "diabetic ketoacidosis", "latent autoimmune diabetes",
        "morbid obesity", "bmi", "bariatric surgery", "weight management",
        "metabolic-associated fatty liver", "mafld", "liver steatosis",
        "liver fibrosis steatosis", "steatohepatitis",
        "lipodystrophy", "hypertriglyceridemia", "dyslipidemia",
        "pheochromocytoma adrenal", "primary hyperaldosteronism",
        "congenital adrenal hyperplasia", "addison disease",
        "hypoparathyroidism", "vitamin d deficiency", "rickets",
        "diabetes insipidus", "siadh", "hypopituitarism",
        "hashimoto", "thyroid nodule", "differentiated thyroid",
        # Drug keywords
        "semaglutide", "liraglutide", "tirzepatide", "dulaglutide",
        "exenatide", "albiglutide", "glp-1",
        "empagliflozin", "dapagliflozin", "canagliflozin", "ertugliflozin",
        "sitagliptin", "saxagliptin", "alogliptin", "linagliptin",
        "insulin degludec", "insulin glargine", "insulin lispro",
        "metformin", "pioglitazone", "resmetirom", "lanifibranor",
        "obeticholic acid", "seladelpar", "elafibranor",
        "evolocumab metabolic", "alirocumab metabolic",
    ],
    "Hematology": [
        # Core conditions
        "sickle cell disease", "hemophilia", "thalassemia",
        "iron deficiency anemia", "aplastic anemia", "myelodysplastic syndrome",
        "essential thrombocythemia", "polycythemia vera",
        "von willebrand disease", "hemolytic anemia",
        "paroxysmal nocturnal hemoglobinuria", "thrombocytopenia", "neutropenia",
        "blood coagulation disorder",
        # Extended conditions
        "beta-thalassemia", "alpha-thalassemia", "sickle cell anemia",
        "immune thrombocytopenia", "itp", "thrombotic thrombocytopenic purpura",
        "ttp", "atypical hemolytic uremic syndrome", "ahus",
        "cold agglutinin disease", "warm autoimmune hemolytic anemia",
        "hemolytic uremic syndrome",
        "red cell aplasia", "pure red cell aplasia", "diamond-blackfan",
        "chronic myeloid", "myeloproliferative neoplasm",
        "bone marrow failure", "hematopoietic stem cell transplant",
        "allogeneic transplant", "autologous transplant",
        "fetal hemoglobin induction", "hemoglobin switching",
        "factor viii", "factor ix", "factor x", "factor xi",
        "deep vein thrombosis hematology", "venous thromboembolism",
        # Drug keywords
        "luspatercept", "voxelotor", "crizanlizumab", "l-glutamine sickle",
        "emicizumab", "fitusiran", "concizumab", "marstacimab",
        "valoctocogene roxaparvovec", "fidanacogene elaparvovec",
        "betibeglogene", "lovotibeglogene",
        "ruxolitinib", "fedratinib", "pacritinib", "momelotinib",
        "avacopan", "iptacopan", "danicopan", "pegcetacoplan", "ravulizumab",
        "eculizumab", "roxadustat", "daprodustat", "vadadustat",
    ],
    "Dermatology": [
        # Core conditions
        "atopic dermatitis", "psoriasis", "acne vulgaris", "alopecia areata",
        "rosacea", "vitiligo", "hidradenitis suppurativa", "chronic urticaria",
        "prurigo nodularis", "ichthyosis", "pemphigus", "pemphigoid",
        "onychomycosis", "wound healing", "keloid",
        # Extended conditions
        "eczema", "hand eczema", "chronic hand dermatitis",
        "plaque psoriasis", "palmoplantar psoriasis", "scalp psoriasis",
        "erythrodermic psoriasis", "pustular psoriasis",
        "bullous pemphigoid", "cicatricial pemphigoid", "pemphigus vulgaris",
        "epidermolysis bullosa", "congenital ichthyosis",
        "pyoderma gangrenosum", "calciphylaxis", "lipodermatosclerosis",
        "subcutaneous fat necrosis", "morphea", "lichen planus",
        "lichen sclerosus", "cutaneous t-cell lymphoma",
        "toxic epidermal necrolysis", "ten", "stevens-johnson syndrome",
        "erythema multiforme", "drug reaction eosinophilia",
        "pityriasis rubra pilaris", "urticarial vasculitis",
        "chronic inducible urticaria", "pruritus", "chronic pruritus",
        # Drug keywords
        "dupilumab derm", "tralokinumab", "lebrikizumab", "nemolizumab",
        "cendakimab", "abrocitinib", "upadacitinib derm",
        "baricitinib derm", "ruxolitinib cream", "tapinarof", "difamilast",
        "secukinumab derm", "guselkumab derm", "risankizumab derm",
        "tildrakizumab", "bimekizumab derm", "spesolimab derm",
        "imsidolimab", "apremilast", "deucravacitinib",
    ],
    "Ophthalmology": [
        # Core conditions
        "age-related macular degeneration", "macular degeneration",
        "diabetic retinopathy", "glaucoma", "dry eye disease", "uveitis",
        "retinal vein occlusion", "corneal disease", "ocular surface disease",
        "geographic atrophy", "neovascular macular", "optic neuritis",
        "inherited retinal dystrophy",
        # Extended conditions
        "wet amd", "neovascular amd", "choroidal neovascularization",
        "diabetic macular edema", "proliferative diabetic retinopathy",
        "branch retinal vein occlusion", "central retinal vein occlusion",
        "retinitis pigmentosa", "leber congenital amaurosis", "choroideremia",
        "stargardt disease", "best disease", "achromatopsia",
        "birdshot chorioretinopathy", "vogt-koyanagi-harada",
        "sympathetic ophthalmia", "multifocal choroiditis",
        "vitreomacular traction", "epiretinal membrane", "macular hole",
        "ocular hypertension", "primary open-angle glaucoma",
        "angle-closure glaucoma", "normal tension glaucoma",
        "keratoconus", "corneal dystrophy", "fuchs endothelial dystrophy",
        "sjogren dry eye", "graft-versus-host ocular",
        "thyroid eye disease", "graves ophthalmopathy",
        "retinal dystrophy", "rod-cone dystrophy",
        # Drug keywords
        "ranibizumab", "bevacizumab eye", "aflibercept", "brolucizumab",
        "faricimab", "avacincaptad pegol", "pegcetacoplan eye",
        "zimura", "izervay", "oracea eye", "lifitegrast", "cyclosporine eye",
        "latanoprost", "bimatoprost", "travoprost", "timolol eye",
        "netarsudil", "omidenepag", "tepoxalin eye",
    ],
    "Gastroenterology & Hepatology": [
        # Core conditions
        "primary biliary cholangitis", "primary sclerosing cholangitis",
        "irritable bowel syndrome", "eosinophilic esophagitis",
        "portal hypertension", "liver cirrhosis", "autoimmune hepatitis",
        "short bowel syndrome", "celiac disease", "functional dyspepsia",
        "gastroparesis", "cholestasis", "biliary atresia",
        # Extended conditions
        "nonalcoholic steatohepatitis", "nash liver", "nafld liver", "mafld",
        "non-alcoholic fatty liver", "metabolic fatty liver",
        "liver steatosis", "hepatic steatosis", "steatohepatitis",
        "microscopic colitis", "collagenous colitis", "lymphocytic colitis",
        "pouchitis", "intestinal failure", "small bowel",
        "helicobacter pylori", "h. pylori", "peptic ulcer disease",
        "gastroesophageal reflux", "gerd", "barrett esophagus",
        "achalasia", "esophageal motility",
        "functional constipation", "chronic constipation", "ibs-c", "ibs-d",
        "hepatic encephalopathy", "esophageal varices", "spontaneous bacterial peritonitis",
        "acute liver failure", "drug-induced liver injury", "dili",
        "alcohol-related liver disease", "alcoholic hepatitis",
        "exocrine pancreatic insufficiency", "pancreatitis", "chronic pancreatitis",
        "cholelithiasis", "gallstone", "cholangitis",
        "liver fibrosis", "liver transplant", "post-transplant liver",
        # Drug keywords
        "vedolizumab", "ustekinumab gi", "risankizumab gi", "mirikizumab",
        "etrasimod gi", "ozanimod gi", "filgotinib gi",
        "obeticholic acid", "seladelpar gi", "elafibranor gi",
        "cilofexor", "lanifibranor", "aramchol", "resmetirom",
        "rifaximin", "lactulose", "terlipressin", "vasopressin gi",
        "mesalamine", "sulfasalazine", "budesonide gi",
    ],
    "Psychiatry & Mental Health": [
        # Core conditions
        "major depressive disorder", "treatment-resistant depression",
        "generalized anxiety disorder", "schizophrenia", "bipolar disorder",
        "attention deficit hyperactivity", "autism spectrum",
        "post-traumatic stress", "obsessive-compulsive disorder",
        "substance use disorder", "alcohol use disorder", "opioid use disorder",
        "anorexia nervosa", "bulimia nervosa", "binge eating disorder",
        "borderline personality",
        # Extended conditions
        "depression", "depressive episode", "dysthymia", "persistent depressive",
        "psychotic depression", "postpartum depression", "peripartum depression",
        "seasonal affective", "mdd", "trd",
        "anxiety disorder", "social anxiety", "panic disorder", "agoraphobia",
        "separation anxiety", "selective mutism",
        "bipolar i disorder", "bipolar ii disorder", "cyclothymia",
        "first-episode psychosis", "schizophrenia spectrum",
        "negative symptoms", "cognitive impairment schizophrenia",
        "clozapine-resistant", "treatment-resistant schizophrenia",
        "adhd", "attention deficit",
        "insomnia disorder", "chronic insomnia", "hypersomnolence",
        "narcolepsy type 1", "idiopathic hypersomnia",
        "alcohol dependence", "opioid dependence", "cocaine use disorder",
        "cannabis use disorder", "nicotine dependence",
        "tardive dyskinesia", "extrapyramidal symptoms",
        "trichotillomania", "body dysmorphic disorder", "hoarding",
        "ptsd", "acute stress disorder",
        # Drug keywords
        "zuranolone", "brexanolone", "esketamine", "ketamine depression",
        "psilocybin", "mdma therapy",
        "aripiprazole", "brexpiprazole", "cariprazine", "lumateperone",
        "risperidone", "olanzapine", "quetiapine", "clozapine",
        "xanomeline", "trospium chloride", "emraclidine",
        "valbenazine", "deutetrabenazine", "ingrezza",
        "buprenorphine", "methadone", "naltrexone", "naloxone",
        "modafinil", "solriamfetol", "pitolisant",
        "vortioxetine", "vilazodone", "levomilnacipran", "agomelatine",
    ],
    "Musculoskeletal": [
        # Core conditions
        "osteoporosis", "osteoarthritis", "juvenile idiopathic arthritis",
        "duchenne muscular dystrophy", "spinal cord injury", "fibromyalgia",
        "becker muscular dystrophy", "limb-girdle muscular dystrophy",
        "osteogenesis imperfecta", "bone loss", "fracture prevention",
        # Extended conditions
        "bone mineral density", "vertebral fracture", "hip fracture",
        "low bone density", "osteopenia", "fragility fracture",
        "knee osteoarthritis", "hip osteoarthritis", "hand osteoarthritis",
        "chondromalacia", "cartilage repair", "joint replacement",
        "myopathy", "inflammatory myopathy", "inclusion body myositis",
        "sarcopenia", "muscle wasting", "cachexia musculoskeletal",
        "gout", "hyperuricemia", "calcium pyrophosphate",
        "spinal stenosis", "lumbar disc herniation", "cervical radiculopathy",
        "degenerative disc disease", "spondylolisthesis",
        "tendinopathy", "rotator cuff", "plantar fasciitis", "achilles",
        "muscular dystrophy", "facioscapulohumeral", "myotonic dystrophy",
        "spinal muscular atrophy type",
        # Drug keywords
        "denosumab", "romosozumab", "teriparatide", "abaloparatide",
        "zoledronic acid", "alendronate", "risedronate", "ibandronate",
        "eteplirsen", "golodirsen", "viltolarsen", "casimersen", "ataluren",
        "deflazacort", "vamorolone",
        "burosumab", "setrusumab", "palovarotene",
        "fasinumab musculoskeletal", "tanezumab musculoskeletal",
        "sprifermin", "lorecivivint",
    ],
    "Rare & Genetic Diseases": [
        # Core conditions
        "lysosomal storage disorder", "gaucher disease", "fabry disease",
        "pompe disease", "niemann-pick disease", "phenylketonuria",
        "wilson disease", "amyloidosis", "transthyretin amyloidosis",
        "fragile x syndrome", "rett syndrome", "angelman syndrome",
        "hereditary angioedema", "alpha-1 antitrypsin deficiency",
        "mucopolysaccharidosis", "methylmalonic acidemia",
        "urea cycle disorder", "congenital disorder",
        # Extended conditions
        "hunter syndrome", "hurler syndrome", "maroteaux-lamy syndrome",
        "sanfilippo syndrome", "morquio syndrome",
        "krabbe disease", "metachromatic leukodystrophy",
        "neuronal ceroid lipofuscinosis", "batten disease",
        "acid sphingomyelinase deficiency", "npc1",
        "propionic acidemia", "isovaleric acidemia", "organic aciduria",
        "ornithine transcarbamylase deficiency", "citrullinemia",
        "maple syrup urine disease", "homocystinuria",
        "biotinidase deficiency", "biotin-thiamine responsive",
        "peroxisomal disorder", "zellweger spectrum",
        "hermansky-pudlak syndrome", "chediak-higashi syndrome",
        "neurofibromatosis", "nf1", "nf2",
        "tuberous sclerosis complex", "vhl disease",
        "achondroplasia", "hypochondroplasia", "skeletal dysplasia",
        "kabuki syndrome", "williams syndrome", "prader-willi",
        "22q11 deletion", "charge syndrome",
        "primary hyperoxaluria", "cystinuria", "cystinosis",
        "hereditary hemorrhagic telangiectasia", "hht",
        "porphyria", "acute intermittent porphyria",
        "inherited retinal disease",
        # Drug keywords
        "migalastat", "agalsidase", "alglucosidase alfa", "idursulfase",
        "laronidase", "elosulfase", "galsulfase", "imiglucerase",
        "velaglucerase", "taliglucerase", "eliglustat", "miglustat",
        "cerliponase alfa", "voretigene neparvovec",
        "patisiran", "inotersen", "eplontersen", "vutrisiran",
        "tafamidis", "acoramidis",
        "givosiran", "lumasiran", "nedosiran",
        "avalglucosidase alfa", "cipaglucosidase",
        "setmelanotide", "diazoxide choline",
        "berotralstat", "garadacimab", "garadacimab hae",
    ],
    "Renal & Urology": [
        # Core conditions
        "chronic kidney disease", "iga nephropathy",
        "focal segmental glomerulosclerosis", "polycystic kidney disease",
        "lupus nephritis", "diabetic nephropathy", "acute kidney injury",
        "membranous nephropathy", "alport syndrome", "overactive bladder",
        "interstitial cystitis", "urinary incontinence",
        # Extended conditions
        "ckd", "end-stage renal disease", "esrd", "dialysis",
        "nephrotic syndrome", "minimal change disease", "nephrotic",
        "c3 glomerulopathy", "dense deposit disease",
        "anti-gbm nephropathy", "fibrillary glomerulonephritis",
        "renal transplant", "kidney transplant", "transplant nephrology",
        "calcineurin inhibitor nephrotoxicity",
        "anca-associated nephritis", "rapidly progressive glomerulonephritis",
        "hyperoxaluria renal", "cystinosis renal",
        "benign prostatic hyperplasia", "bph", "lower urinary tract symptoms",
        "luts", "prostate enlargement", "urinary retention",
        "neurogenic bladder", "detrusor overactivity",
        "urinary tract infection", "recurrent uti",
        "kidney stone", "nephrolithiasis", "cystinuria renal",
        "bladder outlet obstruction",
        # Drug keywords
        "sparsentan", "atrasentan", "bardoxolone methyl", "finerenone",
        "empagliflozin renal", "dapagliflozin renal",
        "belimumab nephritis", "voclosporin", "avacopan renal",
        "iptacopan renal", "pegcetacoplan renal",
        "ravulizumab", "eculizumab",
        "oxybutynin", "mirabegron", "vibegron", "solifenacin", "tolterodine",
        "tamsulosin", "silodosin", "alfuzosin",
        "dutasteride", "finasteride bph", "tadalafil bph",
    ],
    "Women's Health": [
        # Core conditions
        "endometriosis", "uterine fibroids", "polycystic ovary syndrome",
        "premature ovarian insufficiency", "menopause", "preeclampsia",
        "preterm birth", "female infertility", "cervical cancer",
        "ovarian cancer", "endometrial cancer", "vulvar",
        # Extended conditions
        "pcos", "hyperandrogenism", "anovulation",
        "dysmenorrhea", "premenstrual syndrome", "pmdd",
        "abnormal uterine bleeding", "heavy menstrual bleeding",
        "menorrhagia", "metrorrhagia", "amenorrhea",
        "adenomyosis", "uterine leiomyoma", "fibroid",
        "vulvodynia", "vestibulodynia", "vulvar lichen sclerosus",
        "vaginismus", "vaginal atrophy", "genitourinary syndrome menopause",
        "hot flash", "vasomotor symptoms", "perimenopausal",
        "postmenopausal osteoporosis", "hormone replacement",
        "breast cancer hormone", "estrogen receptor positive",
        "gestational diabetes", "preeclampsia", "eclampsia", "hellp",
        "recurrent pregnancy loss", "miscarriage", "pregnancy loss",
        "preterm labor", "cervical insufficiency",
        "female sexual dysfunction",
        "assisted reproduction", "ivf", "ovarian stimulation",
        "progesterone supplementation", "luteal phase defect",
        "hrp breast", "hormonal contraception",
        # Drug keywords
        "linzagolix", "elagolix", "relugolix gynecology", "orilissa",
        "fezolinetant", "elinzanetant",
        "ospemifene", "prasterone vaginal", "intrarosa",
        "progesterone", "micronized progesterone",
        "letrozole fertility", "clomiphene citrate",
        "gonadotropin", "follitropin", "lutropin",
        "medroxyprogesterone", "norethindrone", "levonorgestrel",
        "ulipristal acetate", "vilaprisan",
    ],
    "Vaccines": [
        # Core terms
        "vaccine", "vaccination", "immunization", "prophylactic vaccine",
        "mrna vaccine", "cancer vaccine", "preventive immunization",
        # Extended terms
        "adjuvant vaccine", "antigen presentation", "booster dose",
        "prime-boost", "live attenuated", "inactivated vaccine",
        "subunit vaccine", "virus-like particle", "vlp",
        "dna vaccine", "vector vaccine", "adenoviral vector",
        "lipid nanoparticle vaccine", "monovalent", "bivalent", "pentavalent",
        "pediatric vaccine", "adult immunization", "travel vaccine",
        "herd immunity", "seroconversion", "immunogenicity",
        "prophylaxis infection", "pre-exposure prophylaxis vaccine",
        # Specific vaccine targets
        "hpv vaccine", "pneumococcal vaccine", "meningococcal vaccine",
        "typhoid vaccine", "yellow fever vaccine", "rabies vaccine",
        "zoster vaccine", "shingles vaccine",
        "rsv vaccine", "respiratory syncytial virus vaccine",
        "influenza vaccine", "seasonal influenza", "pandemic influenza",
        "covid-19 vaccine", "sars-cov-2 vaccine",
        "ebola vaccine", "dengue vaccine", "malaria vaccine",
        "tuberculosis vaccine", "hiv vaccine",
        "mrna-1273", "bnt162", "chadox",
    ],
    "Pain": [
        # Core conditions
        "chronic pain", "neuropathic pain", "low back pain",
        "cancer pain", "postoperative pain", "fibromyalgia pain",
        "central sensitization", "pain management",
        # Extended conditions
        "complex regional pain syndrome", "crps",
        "allodynia", "hyperalgesia", "wind-up phenomenon",
        "musculoskeletal pain", "chronic musculoskeletal",
        "visceral pain", "chronic pelvic pain", "abdominal pain syndrome",
        "cluster headache", "trigeminal neuralgia",
        "postherpetic neuralgia", "shingles pain",
        "diabetic peripheral neuropathy pain", "dpnp",
        "chemotherapy-induced neuropathy", "cipn",
        "phantom limb pain", "stump pain",
        "failed back surgery syndrome", "post-surgical pain",
        "neck pain", "shoulder pain", "knee pain chronic",
        "osteoarthritis pain",
        "pain catastrophizing", "central sensitization syndrome",
        "small fiber neuropathy pain",
        # Drug keywords
        "tanezumab", "fasinumab pain", "fulranumab",
        "gabapentin pain", "pregabalin pain",
        "duloxetine pain", "amitriptyline pain", "nortriptyline pain",
        "tapentadol", "buprenorphine pain", "oxycodone", "hydromorphone",
        "low-dose naltrexone", "naltrexone pain",
        "ketamine infusion", "lidocaine infusion pain",
        "capsaicin patch", "qutenza", "lidocaine patch",
        "lacosamide pain", "carbamazepine pain", "oxcarbazepine pain",
    ],
}

# ─── DRUG → THERAPY AREA LOOKUP TABLE ─────────────────────────────────────────
# Secondary classification signal: scanned against the interventions list.
# Catches trials where conditions text is vague (e.g. "Advanced Solid Tumor")
# but the drug name alone is a strong therapy area signal.
DRUG_MAP = {
    # ── Oncology ──────────────────────────────────────────────────────────────
    "pembrolizumab":          ["Oncology"],
    "nivolumab":              ["Oncology"],
    "atezolizumab":           ["Oncology"],
    "durvalumab":             ["Oncology"],
    "avelumab":               ["Oncology"],
    "cemiplimab":             ["Oncology"],
    "ipilimumab":             ["Oncology"],
    "tremelimumab":           ["Oncology"],
    "relatlimab":             ["Oncology"],
    "dostarlimab":            ["Oncology"],
    "trastuzumab":            ["Oncology"],
    "pertuzumab":             ["Oncology"],
    "sacituzumab":            ["Oncology"],
    "enfortumab":             ["Oncology"],
    "mirvetuximab":           ["Oncology"],
    "brentuximab":            ["Oncology"],
    "polatuzumab":            ["Oncology"],
    "inotuzumab":             ["Oncology"],
    "gemtuzumab":             ["Oncology"],
    "bevacizumab":            ["Oncology"],
    "ramucirumab":            ["Oncology"],
    "cetuximab":              ["Oncology"],
    "panitumumab":            ["Oncology"],
    "amivantamab":            ["Oncology"],
    "rituximab":              ["Oncology"],
    "obinutuzumab":           ["Oncology"],
    "daratumumab":            ["Oncology"],
    "isatuximab":             ["Oncology"],
    "teclistamab":            ["Oncology"],
    "talquetamab":            ["Oncology"],
    "elranatamab":            ["Oncology"],
    "linvoseltamab":          ["Oncology"],
    "tarlatamab":             ["Oncology"],
    "venetoclax":             ["Oncology", "Hematology"],
    "ibrutinib":              ["Oncology", "Hematology"],
    "acalabrutinib":          ["Oncology", "Hematology"],
    "zanubrutinib":           ["Oncology", "Hematology"],
    "pirtobrutinib":          ["Oncology", "Hematology"],
    "palbociclib":            ["Oncology"],
    "ribociclib":             ["Oncology"],
    "abemaciclib":            ["Oncology"],
    "olaparib":               ["Oncology"],
    "niraparib":              ["Oncology"],
    "rucaparib":              ["Oncology"],
    "talazoparib":            ["Oncology"],
    "adagrasib":              ["Oncology"],
    "sotorasib":              ["Oncology"],
    "divarasib":              ["Oncology"],
    "osimertinib":            ["Oncology"],
    "lazertinib":             ["Oncology"],
    "erlotinib":              ["Oncology"],
    "gefitinib":              ["Oncology"],
    "alectinib":              ["Oncology"],
    "brigatinib":             ["Oncology"],
    "lorlatinib":             ["Oncology"],
    "crizotinib":             ["Oncology"],
    "cabozantinib":           ["Oncology"],
    "lenvatinib":             ["Oncology"],
    "sorafenib":              ["Oncology"],
    "regorafenib":            ["Oncology"],
    "sunitinib":              ["Oncology"],
    "imatinib":               ["Oncology", "Hematology"],
    "dasatinib":              ["Oncology", "Hematology"],
    "nilotinib":              ["Oncology", "Hematology"],
    "ponatinib":              ["Oncology", "Hematology"],
    "asciminib":              ["Oncology", "Hematology"],
    "midostaurin":            ["Oncology", "Hematology"],
    "gilteritinib":           ["Oncology", "Hematology"],
    "quizartinib":            ["Oncology", "Hematology"],
    "enasidenib":             ["Oncology", "Hematology"],
    "ivosidenib":             ["Oncology", "Hematology"],
    "azacitidine":            ["Oncology", "Hematology"],
    "decitabine":             ["Oncology", "Hematology"],
    "bortezomib":             ["Oncology", "Hematology"],
    "carfilzomib":            ["Oncology", "Hematology"],
    "ixazomib":               ["Oncology", "Hematology"],
    "alpelisib":              ["Oncology"],
    "copanlisib":             ["Oncology", "Hematology"],
    "mogamulizumab":          ["Oncology"],
    "axicabtagene":           ["Oncology", "Hematology"],
    "tisagenlecleucel":       ["Oncology", "Hematology"],
    "ciltacabtagene":         ["Oncology", "Hematology"],
    "idecabtagene":           ["Oncology", "Hematology"],
    "lisocabtagene":          ["Oncology", "Hematology"],
    "brexucabtagene":         ["Oncology", "Hematology"],
    # ── Neurology ─────────────────────────────────────────────────────────────
    "lecanemab":              ["Neurology"],
    "donanemab":              ["Neurology"],
    "aducanumab":             ["Neurology"],
    "gantenerumab":           ["Neurology"],
    "ocrelizumab":            ["Neurology"],
    "ofatumumab":             ["Neurology"],
    "natalizumab":            ["Neurology"],
    "siponimod":              ["Neurology"],
    "ozanimod":               ["Neurology"],
    "ponesimod":              ["Neurology"],
    "cladribine":             ["Neurology"],
    "alemtuzumab":            ["Neurology"],
    "dimethyl fumarate":      ["Neurology"],
    "teriflunomide":          ["Neurology"],
    "fingolimod":             ["Neurology"],
    "erenumab":               ["Neurology"],
    "fremanezumab":           ["Neurology"],
    "galcanezumab":           ["Neurology"],
    "eptinezumab":            ["Neurology"],
    "ubrogepant":             ["Neurology"],
    "rimegepant":             ["Neurology"],
    "atogepant":              ["Neurology"],
    "lasmiditan":             ["Neurology"],
    "prasinezumab":           ["Neurology"],
    "safinamide":             ["Neurology"],
    "opicapone":              ["Neurology"],
    "nusinersen":             ["Neurology"],
    "onasemnogene":           ["Neurology", "Rare & Genetic Diseases"],
    "risdiplam":              ["Neurology", "Rare & Genetic Diseases"],
    # ── Cardiovascular ────────────────────────────────────────────────────────
    "sacubitril":             ["Cardiovascular"],
    "rivaroxaban":            ["Cardiovascular"],
    "apixaban":               ["Cardiovascular"],
    "dabigatran":             ["Cardiovascular"],
    "edoxaban":               ["Cardiovascular"],
    "ticagrelor":             ["Cardiovascular"],
    "clopidogrel":            ["Cardiovascular"],
    "evolocumab":             ["Cardiovascular"],
    "alirocumab":             ["Cardiovascular"],
    "inclisiran":             ["Cardiovascular"],
    "bempedoic acid":         ["Cardiovascular"],
    "icosapentaenoic":        ["Cardiovascular"],
    "tafamidis":              ["Cardiovascular", "Rare & Genetic Diseases"],
    "acoramidis":             ["Cardiovascular", "Rare & Genetic Diseases"],
    "mavacamten":             ["Cardiovascular"],
    "omecamtiv":              ["Cardiovascular"],
    "vericiguat":             ["Cardiovascular"],
    "abelacimab":             ["Cardiovascular"],
    "asundexian":             ["Cardiovascular"],
    "milvexian":              ["Cardiovascular"],
    # ── Immunology ────────────────────────────────────────────────────────────
    "adalimumab":             ["Immunology & Inflammation"],
    "etanercept":             ["Immunology & Inflammation"],
    "infliximab":             ["Immunology & Inflammation"],
    "golimumab":              ["Immunology & Inflammation"],
    "certolizumab":           ["Immunology & Inflammation"],
    "ustekinumab":            ["Immunology & Inflammation"],
    "secukinumab":            ["Immunology & Inflammation"],
    "ixekizumab":             ["Immunology & Inflammation"],
    "bimekizumab":            ["Immunology & Inflammation"],
    "guselkumab":             ["Immunology & Inflammation"],
    "risankizumab":           ["Immunology & Inflammation", "Gastroenterology & Hepatology"],
    "tildrakizumab":          ["Immunology & Inflammation"],
    "abatacept":              ["Immunology & Inflammation"],
    "tocilizumab":            ["Immunology & Inflammation"],
    "sarilumab":              ["Immunology & Inflammation"],
    "baricitinib":            ["Immunology & Inflammation"],
    "upadacitinib":           ["Immunology & Inflammation"],
    "filgotinib":             ["Immunology & Inflammation"],
    "tofacitinib":            ["Immunology & Inflammation"],
    "anifrolumab":            ["Immunology & Inflammation"],
    "belimumab":              ["Immunology & Inflammation"],
    "voclosporin":            ["Immunology & Inflammation", "Renal & Urology"],
    "avacopan":               ["Immunology & Inflammation", "Renal & Urology"],
    "deucravacitinib":        ["Immunology & Inflammation"],
    "spesolimab":             ["Immunology & Inflammation"],
    "vedolizumab":            ["Immunology & Inflammation", "Gastroenterology & Hepatology"],
    "etrasimod":              ["Immunology & Inflammation", "Gastroenterology & Hepatology"],
    "mirikizumab":            ["Immunology & Inflammation", "Gastroenterology & Hepatology"],
    # ── Metabolic & Endocrine ─────────────────────────────────────────────────
    "semaglutide":            ["Metabolic & Endocrine"],
    "liraglutide":            ["Metabolic & Endocrine"],
    "tirzepatide":            ["Metabolic & Endocrine", "Gastroenterology & Hepatology"],
    "dulaglutide":            ["Metabolic & Endocrine"],
    "exenatide":              ["Metabolic & Endocrine"],
    "empagliflozin":          ["Metabolic & Endocrine", "Cardiovascular"],
    "dapagliflozin":          ["Metabolic & Endocrine", "Cardiovascular"],
    "canagliflozin":          ["Metabolic & Endocrine", "Cardiovascular"],
    "sitagliptin":            ["Metabolic & Endocrine"],
    "saxagliptin":            ["Metabolic & Endocrine"],
    "linagliptin":            ["Metabolic & Endocrine"],
    "resmetirom":             ["Metabolic & Endocrine", "Gastroenterology & Hepatology"],
    "lanifibranor":           ["Metabolic & Endocrine", "Gastroenterology & Hepatology"],
    "obeticholic acid":       ["Gastroenterology & Hepatology"],
    "seladelpar":             ["Gastroenterology & Hepatology"],
    # ── Hematology ────────────────────────────────────────────────────────────
    "luspatercept":           ["Hematology"],
    "voxelotor":              ["Hematology"],
    "crizanlizumab":          ["Hematology"],
    "emicizumab":             ["Hematology"],
    "fitusiran":              ["Hematology"],
    "concizumab":             ["Hematology"],
    "valoctocogene":          ["Hematology"],
    "fidanacogene":           ["Hematology"],
    "betibeglogene":          ["Hematology"],
    "ruxolitinib":            ["Hematology", "Oncology"],
    "fedratinib":             ["Hematology", "Oncology"],
    "pacritinib":             ["Hematology", "Oncology"],
    "momelotinib":            ["Hematology", "Oncology"],
    "iptacopan":              ["Hematology"],
    "danicopan":              ["Hematology"],
    "pegcetacoplan":          ["Hematology"],
    "ravulizumab":            ["Hematology"],
    "eculizumab":             ["Hematology"],
    "roxadustat":             ["Hematology", "Renal & Urology"],
    "daprodustat":            ["Hematology", "Renal & Urology"],
    # ── Dermatology ───────────────────────────────────────────────────────────
    "dupilumab":              ["Dermatology", "Respiratory"],
    "tralokinumab":           ["Dermatology"],
    "lebrikizumab":           ["Dermatology"],
    "nemolizumab":            ["Dermatology"],
    "cendakimab":             ["Dermatology"],
    "abrocitinib":            ["Dermatology"],
    "tapinarof":              ["Dermatology"],
    "difamilast":             ["Dermatology"],
    "apremilast":             ["Dermatology"],
    # ── Ophthalmology ─────────────────────────────────────────────────────────
    "ranibizumab":            ["Ophthalmology"],
    "aflibercept":            ["Ophthalmology"],
    "brolucizumab":           ["Ophthalmology"],
    "faricimab":              ["Ophthalmology"],
    "avacincaptad":           ["Ophthalmology"],
    "amsbarkestat":           ["Ophthalmology"],
    "lifitegrast":            ["Ophthalmology"],
    "netarsudil":             ["Ophthalmology"],
    "omidenepag":             ["Ophthalmology"],
    # ── Rare & Genetic Diseases ───────────────────────────────────────────────
    "migalastat":             ["Rare & Genetic Diseases"],
    "agalsidase":             ["Rare & Genetic Diseases"],
    "alglucosidase":          ["Rare & Genetic Diseases"],
    "idursulfase":            ["Rare & Genetic Diseases"],
    "laronidase":             ["Rare & Genetic Diseases"],
    "eliglustat":             ["Rare & Genetic Diseases"],
    "miglustat":              ["Rare & Genetic Diseases"],
    "cerliponase":            ["Rare & Genetic Diseases"],
    "voretigene":             ["Rare & Genetic Diseases", "Ophthalmology"],
    "patisiran":              ["Rare & Genetic Diseases"],
    "inotersen":              ["Rare & Genetic Diseases"],
    "eplontersen":            ["Rare & Genetic Diseases", "Cardiovascular"],
    "vutrisiran":             ["Rare & Genetic Diseases"],
    "givosiran":              ["Rare & Genetic Diseases"],
    "lumasiran":              ["Rare & Genetic Diseases"],
    "avalglucosidase":        ["Rare & Genetic Diseases"],
    "setmelanotide":          ["Rare & Genetic Diseases"],
    "berotralstat":           ["Rare & Genetic Diseases"],
    "garadacimab":            ["Rare & Genetic Diseases"],
    # ── Infectious Disease ────────────────────────────────────────────────────
    "lenacapavir":            ["Infectious Disease"],
    "cabotegravir":           ["Infectious Disease"],
    "dolutegravir":           ["Infectious Disease"],
    "nirmatrelvir":           ["Infectious Disease"],
    "remdesivir":             ["Infectious Disease"],
    "molnupiravir":           ["Infectious Disease"],
    "glecaprevir":            ["Infectious Disease"],
    "pibrentasvir":           ["Infectious Disease"],
    "sofosbuvir":             ["Infectious Disease"],
    "ceftazidime-avibactam":  ["Infectious Disease"],
    "cefiderocol":            ["Infectious Disease"],
    "isavuconazole":          ["Infectious Disease"],
    "voriconazole":           ["Infectious Disease"],
    "ibrexafungerp":          ["Infectious Disease"],
    "olorofim":               ["Infectious Disease"],
    # ── Respiratory ───────────────────────────────────────────────────────────
    "tezepelumab":            ["Respiratory"],
    "mepolizumab":            ["Respiratory"],
    "benralizumab":           ["Respiratory"],
    "omalizumab":             ["Respiratory"],
    "reslizumab":             ["Respiratory"],
    "nintedanib":             ["Respiratory"],
    "pirfenidone":            ["Respiratory"],
    "elexacaftor":            ["Respiratory"],
    "ivacaftor":              ["Respiratory"],
    "tezacaftor":             ["Respiratory"],
    "lumacaftor":             ["Respiratory"],
    "sotatercept":            ["Respiratory", "Cardiovascular"],
    "selexipag":              ["Respiratory", "Cardiovascular"],
    "macitentan":             ["Respiratory", "Cardiovascular"],
    "riociguat":              ["Respiratory", "Cardiovascular"],
    # ── Psychiatry & Mental Health ────────────────────────────────────────────
    "zuranolone":             ["Psychiatry & Mental Health"],
    "brexanolone":            ["Psychiatry & Mental Health"],
    "esketamine":             ["Psychiatry & Mental Health"],
    "psilocybin":             ["Psychiatry & Mental Health"],
    "xanomeline":             ["Psychiatry & Mental Health"],
    "emraclidine":            ["Psychiatry & Mental Health"],
    "valbenazine":            ["Psychiatry & Mental Health"],
    "deutetrabenazine":       ["Psychiatry & Mental Health"],
    "lumateperone":           ["Psychiatry & Mental Health"],
    "cariprazine":            ["Psychiatry & Mental Health"],
    "brexpiprazole":          ["Psychiatry & Mental Health"],
    "solriamfetol":           ["Psychiatry & Mental Health"],
    "pitolisant":             ["Psychiatry & Mental Health"],
    # ── Musculoskeletal ───────────────────────────────────────────────────────
    "denosumab":              ["Musculoskeletal"],
    "romosozumab":            ["Musculoskeletal"],
    "teriparatide":           ["Musculoskeletal"],
    "abaloparatide":          ["Musculoskeletal"],
    "zoledronic acid":        ["Musculoskeletal"],
    "eteplirsen":             ["Musculoskeletal", "Rare & Genetic Diseases"],
    "golodirsen":             ["Musculoskeletal", "Rare & Genetic Diseases"],
    "viltolarsen":            ["Musculoskeletal", "Rare & Genetic Diseases"],
    "casimersen":             ["Musculoskeletal", "Rare & Genetic Diseases"],
    "ataluren":               ["Musculoskeletal", "Rare & Genetic Diseases"],
    "burosumab":              ["Musculoskeletal", "Rare & Genetic Diseases"],
    "sprifermin":             ["Musculoskeletal"],
    # ── Renal & Urology ───────────────────────────────────────────────────────
    "sparsentan":             ["Renal & Urology"],
    "atrasentan":             ["Renal & Urology"],
    "bardoxolone":            ["Renal & Urology"],
    "finerenone":             ["Renal & Urology", "Cardiovascular"],
    "mirabegron":             ["Renal & Urology"],
    "vibegron":               ["Renal & Urology"],
    # ── Women's Health ────────────────────────────────────────────────────────
    "linzagolix":             ["Women's Health"],
    "elagolix":               ["Women's Health"],
    "relugolix":              ["Women's Health"],
    "fezolinetant":           ["Women's Health"],
    "elinzanetant":           ["Women's Health"],
    "ospemifene":             ["Women's Health"],
    "prasterone":             ["Women's Health"],
    "vilaprisan":             ["Women's Health"],
    # ── Pain ──────────────────────────────────────────────────────────────────
    "tanezumab":              ["Pain", "Musculoskeletal"],
    "fasinumab":              ["Pain"],
    "capsaicin patch":        ["Pain"],
}

AREA_COLORS = {
    "Oncology":                    "#e53935",
    "Neurology":                   "#8e24aa",
    "Cardiovascular":              "#d81b60",
    "Immunology & Inflammation":   "#f4511e",
    "Infectious Disease":          "#6d4c41",
    "Respiratory":                 "#0097a7",
    "Metabolic & Endocrine":       "#2e7d32",
    "Hematology":                  "#c62828",
    "Dermatology":                 "#e65100",
    "Ophthalmology":               "#1565c0",
    "Gastroenterology & Hepatology": "#558b2f",
    "Psychiatry & Mental Health":  "#4527a0",
    "Musculoskeletal":             "#546e7a",
    "Rare & Genetic Diseases":     "#ad1457",
    "Renal & Urology":             "#00695c",
    "Women's Health":              "#c2185b",
    "Vaccines":                    "#00838f",
    "Pain":                        "#e65100",
    "Other":                       "#78909c",
}

STATUS_COLORS = {
    "RECRUITING":               "#1b5e20",
    "ACTIVE_NOT_RECRUITING":    "#0d47a1",
    "NOT_YET_RECRUITING":       "#e65100",
    "ENROLLING_BY_INVITATION":  "#4a148c",
}

# ─── HELPERS ───────────────────────────────────────────────────────────────────
def safe_get(d, *keys):
    for k in keys:
        if not isinstance(d, dict):
            return None
        d = d.get(k)
    return d

def classify_therapy_areas(conditions, title, summary, interventions=None):
    """
    Two-pass classification:
    1. Keyword scan: conditions + title + summary[:600] against THERAPY_AREAS
    2. Drug map scan: interventions list against DRUG_MAP (catches vague titles
       like 'Advanced Solid Tumor' where the drug name is the clearest signal)
    Returns a deduplicated, sorted list of matched areas, or ["Other"].
    """
    # ── Pass 1: keyword scan ─────────────────────────────────────────────────
    text = " " + " ".join(filter(None, [
        " ".join(conditions or []),
        title or "",
        (summary or "")[:600],
        " ".join(interventions or []),   # also scan intervention names
    ])).lower() + " "

    matched = set()
    for area, kws in THERAPY_AREAS.items():
        for kw in kws:
            if kw in text:
                matched.add(area)
                break

    # ── Pass 2: drug map scan ────────────────────────────────────────────────
    if interventions:
        intv_text = " ".join(interventions).lower()
        for drug_kw, areas in DRUG_MAP.items():
            if drug_kw in intv_text:
                matched.update(areas)

    return sorted(matched) if matched else ["Other"]

# ─── LLM CLASSIFICATION ──────────────────────────────────────────────────────
def classify_with_claude(records, cache_path="classification_cache.json", api_key=None):
    """
    Classify therapy areas using Claude Haiku with batching and caching.
    Only classifies records not already in cache. Falls back to keywords on error.
    Requires: pip install anthropic>=0.39.0 and ANTHROPIC_API_KEY env var.
    """
    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key)
    except ImportError:
        print("  ⚠  anthropic package not installed. Falling back to keyword classification.")
        print("     To enable LLM classification: pip install anthropic>=0.39.0")
        for r in records:
            if "therapy_areas" not in r or not r["therapy_areas"]:
                r["therapy_areas"] = classify_therapy_areas(
                    r.get("conditions", []), r.get("title", ""), r.get("summary", ""),
                    interventions=r.get("interventions", [])
                )
        return records

    # Load existing cache
    cache = {}
    cache_file = Path(cache_path)
    if cache_file.exists():
        try:
            cache = json.loads(cache_file.read_text())
        except Exception as e:
            print(f"  ⚠  Could not load cache from {cache_path}: {e}")

    # Identify uncached records
    uncached = [r for r in records if r.get("nct_id") not in cache]
    cached_ct = len(records) - len(uncached)
    print(f"  Cache: {cached_ct:,} cached, {len(uncached):,} new records to classify")

    if not uncached:
        print("  ✓ All records already classified (cache hit).")
        for r in records:
            r["therapy_areas"] = cache.get(r["nct_id"], ["Other"])
        return records

    # Valid therapy area names for the prompt
    valid_areas = list(AREA_COLORS.keys())

    # Batch and classify
    BATCH_SIZE = 40
    batches = [uncached[i:i+BATCH_SIZE] for i in range(0, len(uncached), BATCH_SIZE)]
    print(f"  Batching {len(uncached):,} records into {len(batches)} API call(s)…")

    for batch_idx, batch in enumerate(batches, 1):
        print(f"    Batch {batch_idx}/{len(batches)} ({len(batch)} trials)…", end=" ", flush=True)

        trial_list = []
        for r in batch:
            trial_list.append({
                "nct_id":        r.get("nct_id"),
                "title":         r.get("title", ""),
                "conditions":    r.get("conditions", [])[:4],
                "interventions": r.get("interventions", [])[:4],
            })

        prompt = f"""You are a clinical trial classifier. For each trial below, classify it into one or more therapy areas from ONLY this list:
{', '.join(valid_areas)}

Rules:
- Use ONLY areas from the list above (exact spelling).
- A trial can belong to multiple areas.
- If none match clearly, use "Other".
- Return ONLY valid JSON — no markdown, no explanation.

Trials:
{json.dumps(trial_list, separators=(',',':'))}

Return a JSON object mapping each NCT ID to an array of therapy areas. Example:
{{"NCT01234567": ["Oncology"], "NCT09876543": ["Cardiovascular", "Metabolic & Endocrine"]}}"""

        try:
            resp = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=4096,
                messages=[{"role": "user", "content": prompt}]
            )
            text = resp.content[0].text

            # Robust JSON extraction (handle markdown code blocks)
            if "```" in text:
                start = text.find("{")
                end   = text.rfind("}") + 1
                if start >= 0 and end > start:
                    text = text[start:end]

            classifications = json.loads(text)

            # Validate and update cache
            for nct_id, areas in classifications.items():
                # Filter to only valid area names
                valid = [a for a in areas if a in AREA_COLORS]
                cache[nct_id] = valid if valid else ["Other"]

            print("✓")
        except Exception as e:
            print(f"✗ ({e})")
            # Fall back to keyword classification for this batch
            for r in batch:
                kw_areas = classify_therapy_areas(
                    r.get("conditions", []), r.get("title", ""),
                    r.get("summary", ""), r.get("interventions", [])
                )
                cache[r["nct_id"]] = kw_areas

        time.sleep(0.3)  # polite pacing

    # Save updated cache
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    cache_file.write_text(json.dumps(cache, indent=2), encoding="utf-8")
    print(f"  ✓ Cache saved → {cache_path} ({len(cache):,} entries)")

    # Apply all classifications to records
    for r in records:
        r["therapy_areas"] = cache.get(r["nct_id"], ["Other"])

    return records

# ─── FETCH ─────────────────────────────────────────────────────────────────────
def fetch_all_studies(max_pages=None, days=365):
    from datetime import timedelta
    today      = datetime.now()
    since_date = (today - timedelta(days=days)).strftime("%Y-%m-%d")
    until_date = today.strftime("%Y-%m-%d")

    print(f"\n{'='*62}")
    print(f"  ClinicalTrials.gov Fetch  |  {today.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Start date range : {since_date}  →  {until_date}")
    if max_pages is not None:
        print(f"  Page cap         : {max_pages} pages (≤ {max_pages*1000:,} records)")
    else:
        print(f"  Page cap         : unlimited — fetching all matching trials")
    print(f"{'='*62}")

    studies, page_token, page = [], None, 0

    while True:
        if max_pages is not None and page >= max_pages:
            print(f"\n  ⚠  Stopped at max-pages cap ({max_pages}).")
            break

        params = {
            "filter.advanced": (
                f"AREA[StartDate]RANGE[{since_date},{until_date}]"
            ),
            "pageSize": PAGE_SIZE,
        }
        if page_token:
            params["pageToken"] = page_token

        try:
            r = requests.get(API_BASE, params=params, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"\n  ✗ API error (page {page+1}): {e}")
            break

        batch = data.get("studies", [])
        if not batch:
            break

        studies.extend(batch)
        page_token = data.get("nextPageToken")
        total = data.get("totalCount", "?")
        page += 1
        print(f"  Page {page:>3} | +{len(batch):>5,} | Running total: {len(studies):>6,} / {total}", end="\r")

        if not page_token:
            print(f"\n  ✓ Fetch complete — {len(studies):,} studies retrieved.")
            break

        time.sleep(REQUEST_DELAY)

    return studies, since_date, until_date

# ─── PROCESS ───────────────────────────────────────────────────────────────────
def process_studies(raw):
    print(f"\n  Processing {len(raw):,} studies …")
    out = []
    for s in raw:
        ps = s.get("protocolSection", {})

        nct_id   = safe_get(ps, "identificationModule", "nctId") or ""
        title    = safe_get(ps, "identificationModule", "briefTitle") or ""
        status   = safe_get(ps, "statusModule", "overallStatus") or ""
        start    = safe_get(ps, "statusModule", "startDateStruct", "date") or ""
        end      = safe_get(ps, "statusModule", "primaryCompletionDateStruct", "date") or ""
        updated  = safe_get(ps, "statusModule", "lastUpdatePostDateStruct", "date") or ""
        sponsor  = safe_get(ps, "sponsorCollaboratorsModule", "leadSponsor", "name") or ""
        collab_r = safe_get(ps, "sponsorCollaboratorsModule", "collaborators") or []
        collabs  = [c["name"] for c in collab_r if c.get("name")]
        conds    = safe_get(ps, "conditionsModule", "conditions") or []
        phases_r = safe_get(ps, "designModule", "phases") or []
        phase    = ", ".join(p.replace("PHASE", "Phase ").replace("_", "/") for p in phases_r) if phases_r else "N/A"
        enroll   = safe_get(ps, "designModule", "enrollmentInfo", "count") or ""
        summary  = safe_get(ps, "descriptionModule", "briefSummary") or ""
        org_type = safe_get(ps, "sponsorCollaboratorsModule", "leadSponsor", "class") or ""

        intv_r      = safe_get(ps, "armsInterventionsModule", "interventions") or []
        intvs       = [i["name"] for i in intv_r if i.get("name")]
        intv_types  = sorted({i["type"] for i in intv_r if i.get("type")})

        locs    = safe_get(ps, "contactsLocationsModule", "locations") or []
        countries = sorted(set(loc["country"] for loc in locs if loc.get("country")))

        # ── NEW FIELDS ──────────────────────────────────────────────────────────
        study_type     = safe_get(ps, "designModule", "studyType") or ""

        # Funder types: try the explicit fundedBys array first;
        # if absent (common in v2 responses), derive from leadSponsor.class
        # + each collaborator's class — these are the same enum values.
        funded_bys_r   = safe_get(ps, "sponsorCollaboratorsModule", "fundedBys") or []
        funder_types   = [f for f in funded_bys_r if f]
        if not funder_types:
            classes = []
            if org_type:
                classes.append(org_type)
            for c in collab_r:
                cls = c.get("class", "")
                if cls and cls not in classes:
                    classes.append(cls)
            funder_types = classes

        official_title = safe_get(ps, "identificationModule", "officialTitle") or ""
        acronym        = safe_get(ps, "identificationModule", "acronym") or ""

        elig           = safe_get(ps, "eligibilityModule") or {}
        elig_criteria  = elig.get("eligibilityCriteria") or ""
        min_age        = elig.get("minimumAge") or ""
        max_age        = elig.get("maximumAge") or ""
        sex            = elig.get("sex") or ""
        healthy_vols   = elig.get("healthyVolunteers")

        outcomes_m     = safe_get(ps, "outcomesModule") or {}
        primary_outs   = [o.get("measure", "") for o in (outcomes_m.get("primaryOutcomes") or [])[:4] if o.get("measure")]
        secondary_outs = [o.get("measure", "") for o in (outcomes_m.get("secondaryOutcomes") or [])[:3] if o.get("measure")]

        central_ctcts  = safe_get(ps, "contactsLocationsModule", "centralContacts") or []
        contacts       = [{"name": c.get("name",""), "email": c.get("email",""), "phone": c.get("phone","")}
                          for c in central_ctcts[:2] if c.get("name")]

        design_info    = safe_get(ps, "designModule", "designInfo") or {}
        allocation     = design_info.get("allocation") or ""
        int_model      = design_info.get("interventionModel") or ""
        primary_purpose= design_info.get("primaryPurpose") or ""
        masking_info   = design_info.get("maskingInfo") or {}
        masking        = masking_info.get("masking") or ""
        # ───────────────────────────────────────────────────────────────────────

        ta = classify_therapy_areas(conds, title, summary, interventions=intvs)

        out.append({
            "nct_id":               nct_id,
            "title":                title,
            "official_title":       official_title,
            "acronym":              acronym,
            "status":               status,
            "phase":                phase,
            "study_type":           study_type,
            "sponsor":              sponsor,
            "org_type":             org_type,
            "funder_types":         funder_types,
            "collaborators":        collabs[:8],
            "conditions":           conds[:6],
            "interventions":        intvs[:8],
            "intervention_types":   intv_types,
            "countries":            countries,
            "therapy_areas":        ta,
            "start_date":           start,
            "completion_date":      end,
            "last_update":          updated,
            "enrollment":           enrollment if (enrollment := str(enroll)) != "0" else "",
            "summary":              (summary[:600] + "…") if len(summary) > 600 else summary,
            "url":                  f"https://clinicaltrials.gov/study/{nct_id}",
            "eligibility_criteria": (elig_criteria[:800] + "…") if len(elig_criteria) > 800 else elig_criteria,
            "min_age":              min_age,
            "max_age":              max_age,
            "sex":                  sex,
            "healthy_volunteers":   "Yes" if healthy_vols is True else ("No" if healthy_vols is False else str(healthy_vols or "")),
            "primary_outcomes":     primary_outs,
            "secondary_outcomes":   secondary_outs,
            "contacts":             contacts,
            "allocation":           allocation.replace("_", " ").title() if allocation else "",
            "intervention_model":   int_model.replace("_", " ").title() if int_model else "",
            "primary_purpose":      primary_purpose.replace("_", " ").title() if primary_purpose else "",
            "masking":              masking.replace("_", " ").title() if masking else "",
        })

    print(f"  ✓ Processing done — {len(out):,} records ready.")
    return out

# ─── HTML GENERATOR ────────────────────────────────────────────────────────────
def generate_html(records, generated_at, days=365, since_date=None, until_date=None, viewer_mode=False, github_mode=False):
    from datetime import timedelta
    _today = datetime.now()
    if since_date is None:
        since_date = (_today - timedelta(days=days)).strftime("%Y-%m-%d")
    if until_date is None:
        until_date = _today.strftime("%Y-%m-%d")
    date_range_label = f"{since_date} – {until_date}"

    if viewer_mode:
        # Viewer mode: no embedded data — viewer loads JSON file at runtime
        data_js_block = "let ALL_DATA=[];\nlet AREA_COLORS={};\nlet STATUS_COLORS={};"
        area_checkboxes = phase_checkboxes = status_checkboxes = ""
        country_checkboxes = funder_checkboxes = intv_type_checkboxes = study_type_checkboxes = ""
        total = unique_sponsors = unique_countries = 0
        generated_at_display = "No data loaded — drop a trials_data.json file"
    else:
        data_json   = json.dumps(records, separators=(",", ":"))
        colors_json = json.dumps(AREA_COLORS)
        status_json = json.dumps(STATUS_COLORS)
        data_js_block = f"const ALL_DATA={data_json};\nconst AREA_COLORS={colors_json};\nconst STATUS_COLORS={status_json};"
        total = len(records)
        unique_sponsors = len({r["sponsor"] for r in records if r["sponsor"]})
        unique_countries = len({c for r in records for c in r["countries"]})
        generated_at_display = generated_at

    if not viewer_mode:
        # Build unique filter lists
        all_areas        = sorted({a for r in records for a in r["therapy_areas"]})
        all_phases       = sorted({r["phase"] for r in records if r["phase"] and r["phase"] != "N/A"})
        all_statuses     = sorted({r["status"] for r in records if r["status"]})
        all_countries    = sorted({c for r in records for c in r["countries"]})
        all_funder_types  = sorted({ft for r in records for ft in r.get("funder_types", []) if ft})
        all_study_types   = sorted({r.get("study_type", "") for r in records if r.get("study_type")})
        all_intv_types    = sorted({it for r in records for it in r.get("intervention_types", []) if it})

        def checkboxes(items, css_class, label_fn=None):
            return "\n".join(
                f'<label class="cb-label"><input type="checkbox" class="{css_class}" value="{v}"> '
                f'{label_fn(v) if label_fn else v}</label>'
                for v in items
            )

        area_checkboxes       = checkboxes(all_areas,    "area-cb")
        phase_checkboxes      = checkboxes(all_phases,   "phase-cb")
        status_checkboxes     = checkboxes(
            all_statuses, "status-cb",
            label_fn=lambda s: s.replace("_", " ").title()
        )
        country_checkboxes    = checkboxes(all_countries, "country-cb")
        _FUNDER_LABELS = {
            "INDUSTRY":  "Industry",
            "NIH":       "NIH",
            "FED":       "U.S. Fed Gov't (non-NIH)",
            "OTHER_GOV": "Other Gov't",
            "INDIV":     "Individual",
            "NETWORK":   "Network",
            "OTHER":     "Other",
            "UNKNOWN":   "Unknown",
        }
        funder_checkboxes     = checkboxes(all_funder_types, "funder-cb",
                                           label_fn=lambda s: _FUNDER_LABELS.get(s, s.replace("_", " ").title()))
        _INTV_TYPE_LABELS = {
            "DRUG": "Drug", "BIOLOGICAL": "Biological", "DEVICE": "Device",
            "PROCEDURE": "Procedure", "BEHAVIORAL": "Behavioral", "RADIATION": "Radiation",
            "DIETARY_SUPPLEMENT": "Dietary Supplement", "GENETIC": "Genetic",
            "COMBINATION_PRODUCT": "Combination Product", "DIAGNOSTIC_TEST": "Diagnostic Test",
            "OTHER": "Other",
        }
        intv_type_checkboxes  = checkboxes(all_intv_types, "intv-type-cb",
                                           label_fn=lambda s: _INTV_TYPE_LABELS.get(s, s.replace("_", " ").title()))
        study_type_checkboxes = checkboxes(all_study_types, "study-type-cb",
                                           label_fn=lambda s: s.replace("_", " ").title())

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Clinical Trials Lead Gen Dashboard</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root {{
  --bg: #f0f2f5;
  --surface: #ffffff;
  --surface2: #f8fafc;
  --border: #e2e8f0;
  --border2: #cbd5e1;
  --text: #0f172a;
  --text-mid: #334155;
  --text-muted: #718096;
  --accent: #4f46e5;
  --accent-light: #eef2ff;
  --radius: 10px;
  --shadow: 0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
  --shadow-md: 0 4px 6px rgba(0,0,0,.07), 0 2px 4px rgba(0,0,0,.06);
}}
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Inter',system-ui,sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}}
a{{color:var(--accent);text-decoration:none}}a:hover{{text-decoration:underline}}

/* HEADER */
.header{{background:var(--surface);border-bottom:1px solid var(--border);padding:14px 28px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100;box-shadow:var(--shadow)}}
.header h1{{font-size:17px;font-weight:700;color:var(--text);letter-spacing:-.4px;display:flex;align-items:center;gap:8px}}
.header h1 span{{color:var(--accent)}}
.header-badge{{background:var(--accent-light);color:var(--accent);font-size:11px;font-weight:700;padding:2px 8px;border-radius:999px;letter-spacing:.3px}}
.header-meta{{font-size:12px;color:var(--text-muted);text-align:right}}
.header-meta strong{{color:var(--text-mid)}}

/* KPI CARDS */
.kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;padding:8px 28px 0}}
.kpi-card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:8px 14px;box-shadow:var(--shadow);position:relative;overflow:hidden;display:flex;align-items:center;gap:10px}}
.kpi-card::before{{content:'';position:absolute;top:0;left:0;bottom:0;width:3px;background:var(--kpi-color,var(--accent))}}
.kpi-card-body{{padding-left:2px}}
.kpi-label{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.7px;color:var(--text-muted);margin-bottom:1px}}
.kpi-value{{font-size:20px;font-weight:700;color:var(--text);line-height:1;font-variant-numeric:tabular-nums}}
.kpi-sub{{font-size:10px;color:var(--text-muted);margin-top:1px}}

/* ANALYTICS TAB */
.analytics-wrap{{padding:20px 28px 40px}}
.analytics-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}
.chart-card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:18px 20px;box-shadow:var(--shadow)}}
.chart-card.full{{grid-column:1/-1}}
.chart-card.third{{grid-column:span 1}}
.chart-title{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--text-muted);margin-bottom:14px;display:flex;align-items:center;justify-content:space-between}}
.chart-title span{{font-size:11px;font-weight:500;text-transform:none;letter-spacing:0;color:var(--text-muted)}}
.chart-container{{position:relative}}
.chart-container.h240{{height:240px}}
.chart-container.h300{{height:300px}}
.chart-container.h360{{height:360px}}
.chart-container.h420{{height:420px}}
.analytics-summary{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}}
.summary-stat{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:10px 14px;text-align:center}}
.summary-stat-value{{font-size:20px;font-weight:700;color:var(--text);font-variant-numeric:tabular-nums}}
.summary-stat-label{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted);margin-top:2px}}

/* TAB BAR */
.tab-bar{{display:flex;gap:4px;padding:14px 28px 0;border-bottom:1px solid var(--border);background:var(--surface);position:sticky;top:49px;z-index:90}}
.tab{{background:none;border:none;padding:8px 16px;font-size:13px;font-weight:600;color:var(--text-muted);cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-1px;display:flex;align-items:center;gap:6px;transition:all .15s}}
.tab:hover{{color:var(--text)}}
.tab.active{{color:var(--accent);border-bottom-color:var(--accent)}}
.tab-badge{{background:var(--accent);color:#fff;font-size:10px;font-weight:700;padding:1px 6px;border-radius:999px;min-width:18px;text-align:center}}
.tab.active .tab-badge{{background:var(--accent)}}
.tab:not(.active) .tab-badge{{background:#94a3b8}}

/* SPONSOR CARDS */
.sponsors-wrap{{padding:16px 28px 28px}}
.sponsors-toolbar{{display:flex;gap:10px;align-items:center;margin-bottom:14px;flex-wrap:wrap}}
.sponsors-toolbar input{{border:1px solid var(--border);border-radius:6px;padding:7px 10px;font-size:13px;outline:none;width:240px}}
.sponsors-toolbar input:focus{{border-color:var(--accent)}}
.sponsors-count{{font-size:12px;color:var(--text-muted);margin-left:auto}}
.sponsors-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:12px}}
.sponsor-card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:14px 16px;cursor:pointer;transition:all .15s;position:relative}}
.sponsor-card:hover{{border-color:var(--accent);box-shadow:0 0 0 3px rgba(79,70,229,.08);transform:translateY(-1px)}}
.sponsor-card-name{{font-size:13px;font-weight:700;color:var(--text);margin-bottom:4px;line-height:1.3}}
.sponsor-card-meta{{font-size:11px;color:var(--text-muted);margin-bottom:8px}}
.sponsor-card-areas{{display:flex;flex-wrap:wrap;gap:3px}}
.sponsor-trial-count{{position:absolute;top:12px;right:14px;font-size:22px;font-weight:800;color:var(--accent);opacity:.18;font-variant-numeric:tabular-nums;line-height:1}}

/* MODAL */
.modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:500;align-items:center;justify-content:center;padding:20px;backdrop-filter:blur(2px)}}
.modal-overlay.open{{display:flex}}
.modal-box{{background:var(--surface);border-radius:14px;width:100%;max-width:680px;max-height:88vh;overflow-y:auto;box-shadow:0 20px 60px rgba(0,0,0,.2)}}
.modal-header{{padding:20px 24px 16px;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;justify-content:space-between;gap:12px;position:sticky;top:0;background:var(--surface);border-radius:14px 14px 0 0;z-index:1}}
.modal-title{{font-size:15px;font-weight:700;color:var(--text);line-height:1.4}}
.modal-close{{background:var(--bg);border:1px solid var(--border);border-radius:6px;width:28px;height:28px;cursor:pointer;font-size:14px;display:flex;align-items:center;justify-content:center;flex-shrink:0;color:var(--text-muted)}}
.modal-close:hover{{background:var(--border);color:var(--text)}}
.modal-body{{padding:20px 24px}}
.modal-grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}
.modal-field{{display:flex;flex-direction:column;gap:3px}}
.modal-field.full{{grid-column:1/-1}}
.modal-label{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--text-muted)}}
.modal-value{{font-size:13px;color:var(--text);line-height:1.5}}
.modal-footer{{padding:14px 24px;border-top:1px solid var(--border);display:flex;justify-content:space-between;align-items:center}}

/* EXPORT BUTTONS */
.export-row{{display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
.btn-xlsx{{background:#1d7a3a;color:#fff}}
.btn-xlsx:hover{{background:#166830}}
.btn-sponsors{{background:#0369a1;color:#fff}}
.btn-sponsors:hover{{background:#075985}}
.btn-sponsors-wrap{{position:relative;display:inline-flex;align-items:center}}
.btn-sponsors-badge{{position:absolute;top:-6px;right:-6px;background:#f59e0b;color:#fff;font-size:9px;font-weight:700;padding:1px 5px;border-radius:999px;pointer-events:none}}

/* FILTER BAR */
.filter-bar{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);margin:18px 28px;padding:16px 20px;box-shadow:var(--shadow);display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}}
.filter-group{{display:flex;flex-direction:column;gap:4px}}
.filter-group label{{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted)}}
.filter-group input, .filter-group select{{border:1px solid var(--border);border-radius:6px;padding:7px 10px;font-size:13px;color:var(--text);background:var(--surface);outline:none;transition:border .15s}}
.filter-group input:focus, .filter-group select:focus{{border-color:var(--accent);box-shadow:0 0 0 3px rgba(79,70,229,.1)}}
#search-input{{width:260px}}
.filter-group select{{width:170px;cursor:pointer}}
.btn{{border:none;border-radius:6px;padding:8px 16px;font-size:13px;font-weight:600;cursor:pointer;transition:all .15s}}
.btn-clear{{background:#f1f5f9;color:var(--text-muted)}}
.btn-clear:hover{{background:#e2e8f0;color:var(--text)}}
.btn-export{{background:var(--accent);color:#fff}}
.btn-export:hover{{background:#4338ca}}

/* MULTI-SELECT DROPDOWNS (shared pattern for all multi-filters) */
.multi-dd{{position:relative}}
.multi-dd-btn{{border:1px solid var(--border);border-radius:6px;padding:7px 10px;font-size:13px;cursor:pointer;background:var(--surface);width:180px;display:flex;align-items:center;justify-content:space-between;gap:6px;white-space:nowrap;overflow:hidden}}
.multi-dd-btn:hover{{border-color:#a0aec0}}
.multi-dd-btn.active{{border-color:var(--accent);background:var(--accent-light)}}
.multi-dd-btn span.label{{overflow:hidden;text-overflow:ellipsis;flex:1}}
.multi-panel{{display:none;position:absolute;top:calc(100% + 4px);left:0;background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow-md);width:240px;max-height:300px;overflow-y:auto;z-index:200;padding:8px}}
.multi-panel.wide{{width:280px}}
.multi-panel.open{{display:block}}
.multi-search{{width:100%;border:1px solid var(--border);border-radius:5px;padding:5px 8px;font-size:12px;margin-bottom:6px;outline:none}}
.multi-search:focus{{border-color:var(--accent)}}
.cb-label{{display:flex;align-items:center;gap:8px;padding:5px 8px;border-radius:6px;cursor:pointer;font-size:13px}}
.cb-label:hover{{background:var(--accent-light)}}
.cb-label input{{cursor:pointer;accent-color:var(--accent)}}
.panel-actions{{display:flex;gap:8px;padding:8px 8px 4px;border-top:1px solid var(--border);margin-top:4px}}
.panel-actions a{{font-size:12px;color:var(--accent);cursor:pointer}}
/* DATE RANGE */
.date-range{{display:flex;gap:6px;align-items:center}}
.date-range input[type=date]{{border:1px solid var(--border);border-radius:6px;padding:7px 8px;font-size:12px;color:var(--text);background:var(--surface);outline:none;width:138px}}
.date-range input[type=date]:focus{{border-color:var(--accent);box-shadow:0 0 0 3px rgba(79,70,229,.1)}}
.date-range span{{font-size:12px;color:var(--text-muted)}}

/* RESULTS INFO */
.results-info{{padding:0 28px 10px;display:flex;align-items:center;justify-content:space-between}}
.results-info span{{font-size:13px;color:var(--text-muted)}}
.results-info strong{{color:var(--text)}}

/* TABLE — scrolls in its own container so headers stick cleanly */
.table-wrap{{padding:0 28px 28px}}
.table-container{{overflow-x:auto;overflow-y:auto;max-height:calc(100vh - 310px);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow)}}
table{{width:100%;border-collapse:collapse;background:var(--surface)}}
th{{background:#f8fafc;border-bottom:2px solid var(--border);padding:10px 12px;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted);white-space:nowrap;cursor:pointer;user-select:none;position:sticky;top:0;z-index:10}}
th:hover{{background:#f1f5f9;color:var(--text)}}
th.sort-asc::after{{content:" ↑";color:var(--accent)}}
th.sort-desc::after{{content:" ↓";color:var(--accent)}}
td{{padding:10px 12px;border-bottom:1px solid var(--border);vertical-align:top;font-size:13px}}
tr:last-child td{{border-bottom:none}}
tr:hover td{{background:#fafbff}}

/* BADGES */
.badge{{display:inline-block;padding:2px 8px;border-radius:999px;font-size:11px;font-weight:600;white-space:nowrap;margin:2px 2px 2px 0}}
.badge-status{{padding:3px 9px;border-radius:6px;font-size:11px;font-weight:700;letter-spacing:.2px}}
.tag-area{{padding:2px 8px;border-radius:5px;font-size:11px;font-weight:600;margin:2px 2px 2px 0;display:inline-block;color:#fff}}

/* TITLE CELL */
.trial-title{{max-width:320px;line-height:1.4}}
.trial-title a{{color:var(--text);font-weight:500}}
.trial-title a:hover{{color:var(--accent)}}
.nct-link{{font-family:monospace;font-size:12px;color:var(--accent);white-space:nowrap}}

/* INTERVENTION TYPE PILLS */
.intv-type-pill{{display:inline-block;font-size:9px;font-weight:700;letter-spacing:.3px;text-transform:uppercase;padding:1px 6px;border-radius:4px;margin:1px 2px 1px 0;white-space:nowrap}}
.itp-DRUG{{background:#dbeafe;color:#1d4ed8}}
.itp-BIOLOGICAL{{background:#ede9fe;color:#6d28d9}}
.itp-DEVICE{{background:#d1fae5;color:#065f46}}
.itp-PROCEDURE{{background:#fef3c7;color:#92400e}}
.itp-BEHAVIORAL{{background:#e0f2fe;color:#0369a1}}
.itp-RADIATION{{background:#fee2e2;color:#b91c1c}}
.itp-DIETARY_SUPPLEMENT{{background:#ecfccb;color:#3f6212}}
.itp-GENETIC{{background:#eef2ff;color:#3730a3}}
.itp-COMBINATION_PRODUCT{{background:#fce7f3;color:#9d174d}}
.itp-DIAGNOSTIC_TEST{{background:#f1f5f9;color:#475569}}
.itp-OTHER{{background:#f3f4f6;color:#6b7280}}

/* COLLABORATOR LINES in table + sponsor cards */
.sponsor-collabs{{font-size:10px;color:var(--text-muted);margin-top:2px;line-height:1.3;font-weight:400;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:160px}}
.role-badge{{display:inline-block;font-size:9px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;padding:1px 6px;border-radius:4px;margin-right:4px;vertical-align:middle}}
.role-lead{{background:#eef2ff;color:#4f46e5;border:1px solid #c7d2fe}}
.role-collab{{background:#f0fdf4;color:#16a34a;border:1px solid #bbf7d0}}
.role-both{{background:#fff7ed;color:#c2410c;border:1px solid #fed7aa}}

/* PAGINATION */
.pagination{{display:flex;align-items:center;justify-content:center;gap:6px;padding:20px 28px}}
.page-btn{{border:1px solid var(--border);background:var(--surface);border-radius:6px;padding:6px 12px;font-size:13px;cursor:pointer;transition:all .15s;color:var(--text)}}
.page-btn:hover{{border-color:var(--accent);color:var(--accent)}}
.page-btn.active{{background:var(--accent);color:#fff;border-color:var(--accent)}}
.page-btn:disabled{{opacity:.4;cursor:default}}
.page-info{{font-size:13px;color:var(--text-muted);padding:0 8px}}

/* MODAL SECTIONS */
.modal-section-label{{font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.8px;color:var(--accent);margin:16px 0 8px;padding-bottom:4px;border-bottom:1px solid var(--accent-light)}}
.modal-section-label:first-child{{margin-top:0}}

/* DETAIL ROW */
.detail-row td{{background:#f8fafc;padding:12px 20px}}
.detail-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;font-size:12px}}
.detail-grid .dl{{display:flex;flex-direction:column;gap:3px}}
.detail-grid .dl dt{{font-weight:700;text-transform:uppercase;font-size:10px;letter-spacing:.5px;color:var(--text-muted)}}
.detail-grid .dl dd{{color:var(--text)}}
.expand-btn{{font-size:11px;color:var(--text-muted);cursor:pointer;white-space:nowrap}}
.expand-btn:hover{{color:var(--accent)}}

/* SPONSOR HIGHLIGHT */
.sponsor-cell{{font-weight:600;color:var(--text);max-width:160px}}
.org-industry{{color:#1a1a2e}}
.org-nih{{color:#1b5e20}}
.org-other{{color:#3e2723}}

@media(max-width:900px){{
  .kpi-grid{{grid-template-columns:repeat(2,1fr)}}
  .detail-grid{{grid-template-columns:1fr 1fr}}
}}
</style>
</head>
<body>

<!-- HEADER -->
<div class="header">
  <h1>Clinical Trials <span>Lead Gen</span> Dashboard <span class="header-badge">LIVE</span></h1>
  <div class="header-meta">
    <span id="meta-refresh">{generated_at_display}</span> &nbsp;·&nbsp; clinicaltrials.gov API v2
  </div>
</div>

<!-- KPI CARDS -->
<div class="kpi-grid">
  <div class="kpi-card" style="--kpi-color:#4f46e5">
    <div class="kpi-card-body">
      <div class="kpi-label">Active Trials</div>
      <div class="kpi-value" id="kpi-trials">{total:,}</div>
      <div class="kpi-sub">Start date: {date_range_label}</div>
    </div>
  </div>
  <div class="kpi-card" style="--kpi-color:#0ea5e9">
    <div class="kpi-card-body">
      <div class="kpi-label">Unique Sponsors</div>
      <div class="kpi-value" id="kpi-sponsors">{unique_sponsors:,}</div>
      <div class="kpi-sub">Target companies</div>
    </div>
  </div>
  <div class="kpi-card" style="--kpi-color:#10b981">
    <div class="kpi-card-body">
      <div class="kpi-label">Countries</div>
      <div class="kpi-value" id="kpi-countries">{unique_countries:,}</div>
      <div class="kpi-sub">Trial locations</div>
    </div>
  </div>
  <div class="kpi-card" style="--kpi-color:#f59e0b">
    <div class="kpi-card-body">
      <div class="kpi-label">Therapy Areas</div>
      <div class="kpi-value" id="kpi-recruiting">{total:,}</div>
      <div class="kpi-sub" id="kpi-recruiting-sub">Loading…</div>
    </div>
  </div>
</div>

<!-- TAB BAR -->
<div class="tab-bar">
  <button class="tab active" id="tab-trials" onclick="switchTab('trials')">◫ Trials <span class="tab-badge" id="tab-trials-badge">{total:,}</span></button>
  <button class="tab" id="tab-sponsors" onclick="switchTab('sponsors')">◈ Sponsors <span class="tab-badge" id="tab-sponsors-badge">{unique_sponsors:,}</span></button>
  <button class="tab" id="tab-analytics" onclick="switchTab('analytics')">▦ Analytics</button>
</div>

<!-- FILTER BAR -->
<div class="filter-bar">

  <div class="filter-group">
    <label>Search</label>
    <input id="search-input" type="text" placeholder="Sponsor, title, NCT ID, drug…" autocomplete="off">
  </div>

  <div class="filter-group">
    <label>Therapy Area</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="area-btn"><span class="label" id="area-btn-label">All Areas</span><span>▾</span></div>
      <div class="multi-panel wide open-panel" id="area-panel">
        {area_checkboxes}
        <div class="panel-actions"><a id="area-select-all">All</a><a id="area-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Phase</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="phase-btn"><span class="label" id="phase-btn-label">All Phases</span><span>▾</span></div>
      <div class="multi-panel" id="phase-panel">
        {phase_checkboxes}
        <div class="panel-actions"><a id="phase-select-all">All</a><a id="phase-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Status</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="status-btn"><span class="label" id="status-btn-label">All Statuses</span><span>▾</span></div>
      <div class="multi-panel" id="status-panel">
        {status_checkboxes}
        <div class="panel-actions"><a id="status-select-all">All</a><a id="status-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Country</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="country-btn"><span class="label" id="country-btn-label">All Countries</span><span>▾</span></div>
      <div class="multi-panel" id="country-panel">
        <input class="multi-search" id="country-search" type="text" placeholder="Search country…">
        {country_checkboxes}
        <div class="panel-actions"><a id="country-select-all">All</a><a id="country-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Intervention Type</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="intv-type-btn"><span class="label" id="intv-type-btn-label">All Types</span><span>▾</span></div>
      <div class="multi-panel" id="intv-type-panel">
        {intv_type_checkboxes}
        <div class="panel-actions"><a id="intv-type-select-all">All</a><a id="intv-type-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Funder Type</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="funder-btn"><span class="label" id="funder-btn-label">All Funders</span><span>▾</span></div>
      <div class="multi-panel" id="funder-panel">
        {funder_checkboxes}
        <div class="panel-actions"><a id="funder-select-all">All</a><a id="funder-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Study Type</label>
    <div class="multi-dd">
      <div class="multi-dd-btn" id="study-type-btn"><span class="label" id="study-type-btn-label">All Types</span><span>▾</span></div>
      <div class="multi-panel" id="study-type-panel">
        {study_type_checkboxes}
        <div class="panel-actions"><a id="study-type-select-all">All</a><a id="study-type-clear-all">Clear</a></div>
      </div>
    </div>
  </div>

  <div class="filter-group">
    <label>Start Date Range</label>
    <div class="date-range">
      <input type="date" id="date-from" title="From">
      <span>→</span>
      <input type="date" id="date-to" title="To">
    </div>
  </div>

  <div style="display:flex;gap:8px;align-items:flex-end">
    <button class="btn btn-clear" id="btn-clear">✕ Clear</button>
    <div class="export-row">
      <button class="btn btn-export" id="btn-export">↓ CSV</button>
      <button class="btn btn-xlsx" id="btn-export-xlsx">↓ Excel</button>
      <div class="btn-sponsors-wrap" title="Export deduplicated sponsor list for SeamlessAI / lead gen import">
        <button class="btn btn-sponsors" id="btn-export-sponsors">↓ Sponsors</button>
        <span class="btn-sponsors-badge" id="sponsors-export-badge">0</span>
      </div>
    </div>
  </div>
</div>

<!-- RESULTS INFO -->
<div class="results-info">
  <span>Showing <strong id="showing-count">–</strong> of <strong id="total-count">–</strong> trials</span>
  <span id="page-info-top"></span>
</div>

<!-- TRIALS PANEL -->
<div id="panel-trials">
<!-- TABLE -->
<div class="table-wrap"><div class="table-container">
<table id="main-table">
  <thead>
    <tr>
      <th data-col="nct_id">NCT ID</th>
      <th data-col="title">Trial Title</th>
      <th data-col="sponsor">Sponsor</th>
      <th data-col="therapy_areas">Therapy Area</th>
      <th data-col="intervention_types">Int. Type</th>
      <th data-col="phase">Phase</th>
      <th data-col="status">Status</th>
      <th data-col="countries">Countries</th>
      <th data-col="start_date">Start</th>
      <th data-col="completion_date">Est. Completion</th>
      <th data-col="enrollment">Enroll.</th>
    </tr>
  </thead>
  <tbody id="table-body"></tbody>
</table>
</div></div>

<!-- PAGINATION -->
<div class="pagination" id="pagination"></div>
</div><!-- /panel-trials -->

<!-- ANALYTICS PANEL -->
<div id="panel-analytics" style="display:none">
<div class="analytics-wrap">

  <!-- Summary row -->
  <div class="analytics-summary">
    <div class="summary-stat"><div class="summary-stat-value" id="an-total">–</div><div class="summary-stat-label">Trials in view</div></div>
    <div class="summary-stat"><div class="summary-stat-value" id="an-recruiting">–</div><div class="summary-stat-label">Recruiting</div></div>
    <div class="summary-stat"><div class="summary-stat-value" id="an-sponsors">–</div><div class="summary-stat-label">Unique Sponsors</div></div>
    <div class="summary-stat"><div class="summary-stat-value" id="an-countries">–</div><div class="summary-stat-label">Countries</div></div>
    <div class="summary-stat"><div class="summary-stat-value" id="an-enrollment">–</div><div class="summary-stat-label">Total Enrollment</div></div>
  </div>

  <!-- Charts grid -->
  <div class="analytics-grid">

    <!-- Row 1: Monthly trend full width -->
    <div class="chart-card full">
      <div class="chart-title">Trials Started per Month <span>based on start date</span></div>
      <div class="chart-container h240"><canvas id="ch-trend"></canvas></div>
    </div>

    <!-- Row 2: Therapy area + Phase -->
    <div class="chart-card">
      <div class="chart-title">Trials by Therapy Area <span>top 15</span></div>
      <div class="chart-container h360"><canvas id="ch-area"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Trials by Phase</div>
      <div class="chart-container h360"><canvas id="ch-phase"></canvas></div>
    </div>

    <!-- Row 3: Status + Funder Type -->
    <div class="chart-card">
      <div class="chart-title">Trials by Status</div>
      <div class="chart-container h300"><canvas id="ch-status"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Trials by Funder Type</div>
      <div class="chart-container h300"><canvas id="ch-funder"></canvas></div>
    </div>

    <!-- Row 4: Intervention types + Study type -->
    <div class="chart-card">
      <div class="chart-title">Trials by Intervention Type</div>
      <div class="chart-container h300"><canvas id="ch-intv-type"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Trials by Study Type</div>
      <div class="chart-container h300"><canvas id="ch-study-type"></canvas></div>
    </div>

    <!-- Row 5: Top countries + Top sponsors full width each -->
    <div class="chart-card">
      <div class="chart-title">Top Countries <span>by trial count</span></div>
      <div class="chart-container h420"><canvas id="ch-country"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Top Lead Sponsors <span>by trial count</span></div>
      <div class="chart-container h420"><canvas id="ch-sponsor"></canvas></div>
    </div>

  </div>
</div>
</div><!-- /panel-analytics -->

<!-- SPONSORS PANEL -->
<div id="panel-sponsors" style="display:none">
  <div class="sponsors-wrap">
    <div class="sponsors-toolbar">
      <input id="sponsor-search" type="text" placeholder="Search sponsors…" oninput="renderSponsors()">
      <div class="filter-group" style="flex-direction:row;align-items:center;gap:6px">
        <label style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--text-muted)">Filter by area:</label>
        <div class="multi-dd">
          <div class="multi-dd-btn" id="sp-area-btn" style="width:160px"><span class="label" id="sp-area-label">All Areas</span><span>▾</span></div>
          <div class="multi-panel wide" id="sp-area-panel">
            {area_checkboxes.replace('class="area-cb"', 'class="sp-area-cb"')}
            <div class="panel-actions"><a id="sp-area-all">All</a><a id="sp-area-clear">Clear</a></div>
          </div>
        </div>
      </div>
      <div class="sponsors-count"><span id="sponsors-showing">–</span> sponsors</div>
    </div>
    <div class="sponsors-grid" id="sponsors-grid"></div>
  </div>
</div>

<!-- MODAL -->
<div class="modal-overlay" id="modal-overlay" onclick="closeModal(event)">
  <div class="modal-box" onclick="event.stopPropagation()">
    <div class="modal-header">
      <div class="modal-title" id="modal-title">Trial Details</div>
      <button class="modal-close" onclick="closeModal()">✕</button>
    </div>
    <div class="modal-body" id="modal-body"></div>
    <div class="modal-footer">
      <a id="modal-ct-link" href="#" target="_blank" style="font-size:12px;font-weight:600">↗ View on ClinicalTrials.gov</a>
      <button class="btn btn-clear" onclick="closeModal()">Close</button>
    </div>
  </div>
</div>

<script>
{data_js_block}

const PAGE_SIZE_UI = 75;
let filtered = [...ALL_DATA];
let currentPage = 1;
let sortCol = "start_date";
let sortDir = "desc";

// ── FILTER STATE ──────────────────────────────────────────────────────────────
const state = {{
  search:     "",
  areas:      new Set(),   // empty = all
  phases:     new Set(),
  statuses:   new Set(),
  countries:  new Set(),
  funders:    new Set(),
  studyTypes: new Set(),
  intTypes:   new Set(),
  dateFrom:   "",
  dateTo:     "",
}};

// Normalise a partial date string (YYYY-MM or YYYY) to YYYY-MM-DD for comparison
function toFullDate(d, endOfPeriod=false) {{
  if (!d) return "";
  if (/^\d{{4}}$/.test(d))       return endOfPeriod ? d + "-12-31" : d + "-01-01";
  if (/^\d{{4}}-\d{{2}}$/.test(d)) return endOfPeriod ? d + "-28"   : d + "-01";
  return d;
}}

// ── FILTER LOGIC ──────────────────────────────────────────────────────────────
function applyFilters() {{
  const q = state.search.toLowerCase();
  filtered = ALL_DATA.filter(r => {{
    if (q && ![r.nct_id, r.title, r.sponsor, ...(r.collaborators||[]), ...r.interventions, ...r.conditions]
              .join(" ").toLowerCase().includes(q)) return false;
    if (state.areas.size      > 0 && !r.therapy_areas.some(a => state.areas.has(a)))           return false;
    if (state.phases.size     > 0 && !state.phases.has(r.phase))                               return false;
    if (state.statuses.size   > 0 && !state.statuses.has(r.status))                            return false;
    if (state.countries.size  > 0 && !r.countries.some(c => state.countries.has(c)))           return false;
    if (state.funders.size    > 0 && !(r.funder_types||[]).some(f => state.funders.has(f)))          return false;
    if (state.studyTypes.size > 0 && !state.studyTypes.has(r.study_type))                            return false;
    if (state.intTypes.size   > 0 && !(r.intervention_types||[]).some(t => state.intTypes.has(t)))   return false;
    if (state.dateFrom) {{
      const rd = toFullDate(r.start_date);
      if (!rd || rd < state.dateFrom) return false;
    }}
    if (state.dateTo) {{
      const rd = toFullDate(r.start_date, true);
      if (!rd || rd > state.dateTo) return false;
    }}
    return true;
  }});

  // Sort
  filtered.sort((a, b) => {{
    let va = a[sortCol], vb = b[sortCol];
    if (Array.isArray(va)) va = va.join(", ");
    if (Array.isArray(vb)) vb = vb.join(", ");
    va = va ?? ""; vb = vb ?? "";
    const cmp = String(va).localeCompare(String(vb), undefined, {{numeric: true}});
    return sortDir === "asc" ? cmp : -cmp;
  }});

  currentPage = 1;
  updateKPIs();
  updateSponsorsBadge();
  renderTable();
  renderPagination();
  if (document.getElementById("tab-analytics").classList.contains("active")) renderAnalytics();
}}

// ── RENDER KPIs (dynamic — reflect current filtered set) ─────────────────────
function updateKPIs() {{
  document.getElementById("kpi-trials").textContent    = filtered.length.toLocaleString();
  document.getElementById("kpi-sponsors").textContent  = new Set(filtered.map(r=>r.sponsor)).size.toLocaleString();
  document.getElementById("kpi-countries").textContent = new Set(filtered.flatMap(r=>r.countries)).size.toLocaleString();
  document.getElementById("showing-count").textContent = filtered.length.toLocaleString();
  document.getElementById("total-count").textContent   = ALL_DATA.length.toLocaleString();
  document.getElementById("tab-trials-badge").textContent = filtered.length.toLocaleString();
  const recruiting = filtered.filter(r => r.status === "RECRUITING").length;
  document.getElementById("kpi-recruiting").textContent = recruiting.toLocaleString();
  document.getElementById("kpi-recruiting-sub").textContent = "Actively recruiting";
}}

// ── BADGE HELPERS ─────────────────────────────────────────────────────────────
function areaTag(area) {{
  const c = AREA_COLORS[area] || "#78909c";
  return `<span class="tag-area" style="background:${{c}}">${{area}}</span>`;
}}

function statusBadge(s) {{
  const c  = STATUS_COLORS[s] || "#78909c";
  const lb = s.replace(/_/g," ").replace(/\b\w/g,x=>x.toUpperCase());
  return `<span class="badge-status" style="background:${{c}}1a;color:${{c}};border:1px solid ${{c}}50">${{lb}}</span>`;
}}

const _INTV_TYPE_LABELS_JS = {{DRUG:"Drug",BIOLOGICAL:"Biological",DEVICE:"Device",PROCEDURE:"Procedure",BEHAVIORAL:"Behavioral",RADIATION:"Radiation",DIETARY_SUPPLEMENT:"Dietary Supplement",GENETIC:"Genetic",COMBINATION_PRODUCT:"Combination Product",DIAGNOSTIC_TEST:"Diagnostic Test",OTHER:"Other"}};
function intvTypePills(types) {{
  return (types||[]).map(t =>
    `<span class="intv-type-pill itp-${{t}}">${{_INTV_TYPE_LABELS_JS[t]||t}}</span>`
  ).join("");
}}

function phaseColor(p) {{
  if (p.includes("4")) return "#0d47a1";
  if (p.includes("3")) return "#1565c0";
  if (p.includes("2")) return "#1976d2";
  if (p.includes("1")) return "#42a5f5";
  return "#90a4ae";
}}

// ── TABLE RENDER ──────────────────────────────────────────────────────────────
function renderTable() {{
  const tbody = document.getElementById("table-body");
  const start = (currentPage - 1) * PAGE_SIZE_UI;
  const slice = filtered.slice(start, start + PAGE_SIZE_UI);

  if (!slice.length) {{
    tbody.innerHTML = `<tr><td colspan="11" style="text-align:center;padding:40px;color:#718096">
      No trials match the current filters.</td></tr>`;
    return;
  }}

  tbody.innerHTML = slice.map((r, i) => {{
    const idx = start + i;
    const areas   = r.therapy_areas.map(areaTag).join("");
    const statusB = statusBadge(r.status);
    const phaseC  = phaseColor(r.phase);
    const phaseBadge = r.phase !== "N/A"
      ? `<span style="color:${{phaseC}};font-weight:700">${{r.phase}}</span>`
      : `<span style="color:#a0aec0">N/A</span>`;
    const countries = r.countries.slice(0,3).join(", ") + (r.countries.length > 3 ? ` +${{r.countries.length-3}}` : "");
    const orgClass  = r.org_type === "INDUSTRY" ? "org-industry" : r.org_type === "NIH" ? "org-nih" : "org-other";

    return `<tr data-idx="${{idx}}" onclick="showModal(${{idx}})" style="cursor:pointer" title="Click for details">
      <td onclick="event.stopPropagation()"><a class="nct-link" href="${{r.url}}" target="_blank">${{r.nct_id}}</a></td>
      <td class="trial-title"><span style="color:var(--text);font-weight:500">${{
        r.title.length > 88 ? r.title.slice(0,88)+"…" : r.title
      }}</span></td>
      <td class="sponsor-cell ${{orgClass}}" title="${{[r.sponsor, ...(r.collaborators||[])].join(' | ')}}">
        <div>${{r.sponsor.length > 28 ? r.sponsor.slice(0,28)+"…" : r.sponsor}}</div>
        ${{(r.collaborators||[]).length ? `<div class="sponsor-collabs" title="${{(r.collaborators||[]).join(', ')}}">${{
          (r.collaborators||[]).slice(0,2).map(c => c.length > 22 ? c.slice(0,22)+"…" : c).join(" · ")
          + ((r.collaborators||[]).length > 2 ? ` +${{(r.collaborators||[]).length - 2}}` : "")
        }}</div>` : ""}}
      }}</td>
      <td>${{areas}}</td>
      <td>${{intvTypePills(r.intervention_types)}}</td>
      <td>${{phaseBadge}}</td>
      <td>${{statusB}}</td>
      <td style="font-size:12px;color:var(--text-mid)">${{countries}}</td>
      <td style="white-space:nowrap;font-size:12px;font-family:'JetBrains Mono',monospace">${{r.start_date || "–"}}</td>
      <td style="white-space:nowrap;font-size:12px;font-family:'JetBrains Mono',monospace">${{r.completion_date || "–"}}</td>
      <td style="text-align:right;font-size:12px;color:var(--text-mid);font-variant-numeric:tabular-nums">${{r.enrollment ? Number(r.enrollment).toLocaleString() : "–"}}</td>
    </tr>`;
  }}).join("");
}}

// ── MODAL ─────────────────────────────────────────────────────────────────────
function showModal(idx) {{
  const r = filtered[idx];
  document.getElementById("modal-title").textContent = r.title;
  document.getElementById("modal-ct-link").href = r.url;
  document.getElementById("modal-body").innerHTML = `
    <div class="modal-section-label">Overview</div>
    <div class="modal-grid">
      <div class="modal-field"><div class="modal-label">NCT ID</div>
        <div class="modal-value"><a href="${{r.url}}" target="_blank" style="font-family:'JetBrains Mono',monospace">${{r.nct_id}}</a></div></div>
      <div class="modal-field"><div class="modal-label">Study Type</div>
        <div class="modal-value">${{r.study_type ? r.study_type.replace(/_/g," ") : "–"}}</div></div>
      <div class="modal-field"><div class="modal-label">Lead Sponsor</div>
        <div class="modal-value" style="font-weight:600">
          <span class="role-badge role-lead">Lead</span>
          <span style="cursor:pointer;color:var(--accent)" onclick="event.stopPropagation();filterBySponsor('${{r.sponsor.replace(/'/g,"\\'")}}');document.getElementById('modal-overlay').style.display='none'">${{r.sponsor || "–"}}</span>
        </div></div>
      ${{(r.collaborators||[]).length ? `
      <div class="modal-field full"><div class="modal-label">Collaborators</div>
        <div class="modal-value">${{(r.collaborators||[]).map(c =>
          `<span class="role-badge role-collab">Collab</span><span style="cursor:pointer;color:var(--accent);margin-right:12px" onclick="event.stopPropagation();filterBySponsor('${{c.replace(/'/g,"\\'")}}');document.getElementById('modal-overlay').style.display='none'">${{c}}</span>`
        ).join("")}}</div></div>` : ""}}
      <div class="modal-field"><div class="modal-label">Funder Type</div>
        <div class="modal-value">${{(r.funder_types||[]).map(f=>({{INDUSTRY:"Industry",NIH:"NIH",FED:"U.S. Fed Gov't (non-NIH)",OTHER_GOV:"Other Gov't",INDIV:"Individual",NETWORK:"Network",OTHER:"Other",UNKNOWN:"Unknown"}}[f]||f.replace(/_/g," "))).join(", ") || "–"}}</div></div>
      <div class="modal-field"><div class="modal-label">Phase</div>
        <div class="modal-value">${{r.phase || "–"}}</div></div>
      <div class="modal-field"><div class="modal-label">Status</div>
        <div class="modal-value">${{statusBadge(r.status)}}</div></div>
      <div class="modal-field"><div class="modal-label">Enrollment</div>
        <div class="modal-value">${{r.enrollment ? Number(r.enrollment).toLocaleString() + " participants" : "–"}}</div></div>
      <div class="modal-field"><div class="modal-label">Last Updated</div>
        <div class="modal-value">${{r.last_update || "–"}}</div></div>
      <div class="modal-field"><div class="modal-label">Start Date</div>
        <div class="modal-value">${{r.start_date || "–"}}</div></div>
      <div class="modal-field"><div class="modal-label">Est. Completion</div>
        <div class="modal-value">${{r.completion_date || "–"}}</div></div>
      ${{r.acronym ? `<div class="modal-field"><div class="modal-label">Acronym</div>
        <div class="modal-value" style="font-family:'JetBrains Mono',monospace;font-weight:600">${{r.acronym}}</div></div>` : ""}}
    </div>

    ${{(r.official_title && r.official_title !== r.title) ? `
    <div class="modal-section-label">Official Title</div>
    <div style="font-size:13px;color:var(--text-mid);line-height:1.6;margin-bottom:14px">${{r.official_title}}</div>` : ""}}

    <div class="modal-section-label">Disease &amp; Therapy</div>
    <div class="modal-grid">
      <div class="modal-field full"><div class="modal-label">Therapy Areas</div>
        <div class="modal-value">${{r.therapy_areas.map(areaTag).join(" ")}}</div></div>
      <div class="modal-field full"><div class="modal-label">Conditions</div>
        <div class="modal-value">${{r.conditions.join(" · ") || "–"}}</div></div>
      ${{r.interventions.length ? `<div class="modal-field full"><div class="modal-label">Interventions</div>
        <div class="modal-value">
          ${{(r.intervention_types||[]).length ? `<div style="margin-bottom:4px">${{intvTypePills(r.intervention_types)}}</div>` : ""}}
          ${{r.interventions.join(" · ")}}
        </div></div>` : ""}}
      ${{""/* collaborators moved to Overview section */}}
    </div>

    ${{(r.allocation || r.intervention_model || r.primary_purpose || r.masking) ? `
    <div class="modal-section-label">Study Design</div>
    <div class="modal-grid">
      ${{r.allocation ? `<div class="modal-field"><div class="modal-label">Allocation</div>
        <div class="modal-value">${{r.allocation}}</div></div>` : ""}}
      ${{r.intervention_model ? `<div class="modal-field"><div class="modal-label">Intervention Model</div>
        <div class="modal-value">${{r.intervention_model}}</div></div>` : ""}}
      ${{r.primary_purpose ? `<div class="modal-field"><div class="modal-label">Primary Purpose</div>
        <div class="modal-value">${{r.primary_purpose}}</div></div>` : ""}}
      ${{r.masking ? `<div class="modal-field"><div class="modal-label">Masking</div>
        <div class="modal-value">${{r.masking}}</div></div>` : ""}}
    </div>` : ""}}

    ${{(r.min_age || r.max_age || r.sex || r.eligibility_criteria) ? `
    <div class="modal-section-label">Eligibility</div>
    <div class="modal-grid">
      ${{r.sex ? `<div class="modal-field"><div class="modal-label">Sex</div>
        <div class="modal-value">${{r.sex.replace(/_/g," ")}}</div></div>` : ""}}
      ${{r.min_age ? `<div class="modal-field"><div class="modal-label">Min Age</div>
        <div class="modal-value">${{r.min_age}}</div></div>` : ""}}
      ${{r.max_age ? `<div class="modal-field"><div class="modal-label">Max Age</div>
        <div class="modal-value">${{r.max_age}}</div></div>` : ""}}
      ${{r.healthy_volunteers ? `<div class="modal-field"><div class="modal-label">Healthy Volunteers</div>
        <div class="modal-value">${{r.healthy_volunteers}}</div></div>` : ""}}
      ${{r.eligibility_criteria ? `<div class="modal-field full"><div class="modal-label">Eligibility Criteria</div>
        <div class="modal-value" style="font-size:12px;line-height:1.7;color:var(--text-mid);white-space:pre-wrap;max-height:180px;overflow-y:auto;border:1px solid var(--border);border-radius:6px;padding:8px 10px">${{r.eligibility_criteria}}</div></div>` : ""}}
    </div>` : ""}}

    ${{(r.primary_outcomes.length || r.secondary_outcomes.length) ? `
    <div class="modal-section-label">Outcome Measures</div>
    <div class="modal-grid">
      ${{r.primary_outcomes.length ? `<div class="modal-field full"><div class="modal-label">Primary Outcomes</div>
        <div class="modal-value"><ul style="margin:0;padding-left:16px;font-size:12px;color:var(--text-mid);line-height:1.8">${{r.primary_outcomes.map(o=>`<li>${{o}}</li>`).join("")}}</ul></div></div>` : ""}}
      ${{r.secondary_outcomes.length ? `<div class="modal-field full"><div class="modal-label">Secondary Outcomes</div>
        <div class="modal-value"><ul style="margin:0;padding-left:16px;font-size:12px;color:var(--text-mid);line-height:1.8">${{r.secondary_outcomes.map(o=>`<li>${{o}}</li>`).join("")}}</ul></div></div>` : ""}}
    </div>` : ""}}

    <div class="modal-section-label">Locations &amp; Contact</div>
    <div class="modal-grid">
      <div class="modal-field full"><div class="modal-label">Countries</div>
        <div class="modal-value">${{r.countries.join(", ") || "–"}}</div></div>
      ${{r.contacts && r.contacts.length ? `<div class="modal-field full"><div class="modal-label">Central Contact</div>
        <div class="modal-value">${{r.contacts.map(c =>
          `<span style="font-weight:600">${{c.name}}</span>`
          + (c.email ? ` &nbsp;·&nbsp; <a href="mailto:${{c.email}}">${{c.email}}</a>` : "")
          + (c.phone ? ` &nbsp;·&nbsp; ${{c.phone}}` : "")).join("<br>")}}</div></div>` : ""}}
    </div>

    ${{r.summary ? `
    <div class="modal-section-label">Brief Summary</div>
    <div style="font-size:13px;line-height:1.8;color:var(--text-mid)">${{r.summary}}</div>` : ""}}
  `;
  document.getElementById("modal-overlay").classList.add("open");
  document.body.style.overflow = "hidden";
}}

function closeModal(e) {{
  if (e && e.target !== document.getElementById("modal-overlay") && e.type !== "click") return;
  document.getElementById("modal-overlay").classList.remove("open");
  document.body.style.overflow = "";
}}
document.addEventListener("keydown", e => {{ if (e.key === "Escape") closeModal({{type:"click"}}); }});

// ── PAGINATION ────────────────────────────────────────────────────────────────
function renderPagination() {{
  const totalPages = Math.ceil(filtered.length / PAGE_SIZE_UI);
  const pg = document.getElementById("pagination");

  if (totalPages <= 1) {{ pg.innerHTML = ""; return; }}

  let html = `<button class="page-btn" onclick="goPage(${{currentPage-1}})" ${{currentPage===1?"disabled":""}}>‹ Prev</button>`;

  const win = 2;
  for (let p = 1; p <= totalPages; p++) {{
    if (p === 1 || p === totalPages || (p >= currentPage-win && p <= currentPage+win)) {{
      html += `<button class="page-btn ${{p===currentPage?"active":""}}" onclick="goPage(${{p}})">${{p}}</button>`;
    }} else if (p === currentPage-win-1 || p === currentPage+win+1) {{
      html += `<span style="padding:0 4px;color:#a0aec0">…</span>`;
    }}
  }}

  html += `<button class="page-btn" onclick="goPage(${{currentPage+1}})" ${{currentPage===totalPages?"disabled":""}}>Next ›</button>`;
  html += `<span class="page-info">Page ${{currentPage}} of ${{totalPages}}</span>`;
  pg.innerHTML = html;
}}

function goPage(p) {{
  const totalPages = Math.ceil(filtered.length / PAGE_SIZE_UI);
  currentPage = Math.max(1, Math.min(totalPages, p));
  renderTable();
  renderPagination();
  window.scrollTo({{top: 0, behavior: "smooth"}});
}}

// ── SORT ──────────────────────────────────────────────────────────────────────
document.querySelectorAll("th[data-col]").forEach(th => {{
  th.addEventListener("click", () => {{
    const col = th.dataset.col;
    if (sortCol === col) sortDir = sortDir === "asc" ? "desc" : "asc";
    else {{ sortCol = col; sortDir = "asc"; }}
    document.querySelectorAll("th").forEach(h => h.classList.remove("sort-asc","sort-desc"));
    th.classList.add(sortDir === "asc" ? "sort-asc" : "sort-desc");
    applyFilters();
  }});
}});

// ── GENERIC MULTI-SELECT DROPDOWN FACTORY ────────────────────────────────────
function makeMultiDD({{btnId, panelId, cbClass, stateKey, allLabel, btnLabelId,
                       selectAllId, clearAllId, searchId=null}}) {{
  const btn   = document.getElementById(btnId);
  const panel = document.getElementById(panelId);

  btn.addEventListener("click", e => {{
    e.stopPropagation();
    document.querySelectorAll(".multi-panel.open").forEach(p => {{ if (p !== panel) p.classList.remove("open"); }});
    panel.classList.toggle("open");
  }});

  panel.addEventListener("click", e => e.stopPropagation());

  // Optional search within panel
  if (searchId) {{
    document.getElementById(searchId).addEventListener("input", e => {{
      const q = e.target.value.toLowerCase();
      panel.querySelectorAll(".cb-label").forEach(l => {{
        l.style.display = l.textContent.toLowerCase().includes(q) ? "" : "none";
      }});
    }});
  }}

  panel.querySelectorAll("." + cbClass).forEach(cb => {{
    cb.addEventListener("change", () => {{
      if (cb.checked) state[stateKey].add(cb.value);
      else            state[stateKey].delete(cb.value);
      updateBtnLabel(); applyFilters();
    }});
  }});

  document.getElementById(selectAllId).addEventListener("click", () => {{
    panel.querySelectorAll("." + cbClass).forEach(cb => cb.checked = true);
    state[stateKey] = new Set([...panel.querySelectorAll("." + cbClass)].map(c=>c.value));
    updateBtnLabel(); applyFilters();
  }});

  document.getElementById(clearAllId).addEventListener("click", () => {{
    panel.querySelectorAll("." + cbClass).forEach(cb => cb.checked = false);
    state[stateKey].clear(); updateBtnLabel(); applyFilters();
  }});

  function updateBtnLabel() {{
    const n = state[stateKey].size;
    const lbl = document.getElementById(btnLabelId);
    lbl.textContent = n === 0 ? allLabel : n === 1 ? [...state[stateKey]][0] : `${{n}} selected`;
    btn.classList.toggle("active", n > 0);
  }}

  return {{ reset: () => {{
    panel.querySelectorAll("." + cbClass).forEach(cb => cb.checked = false);
    state[stateKey].clear(); updateBtnLabel();
  }} }};
}}

// Wire up all four multi-select dropdowns
const areaDD      = makeMultiDD({{ btnId:"area-btn",       panelId:"area-panel",       cbClass:"area-cb",       stateKey:"areas",      allLabel:"All Areas",      btnLabelId:"area-btn-label",       selectAllId:"area-select-all",       clearAllId:"area-clear-all" }});
const phaseDD     = makeMultiDD({{ btnId:"phase-btn",      panelId:"phase-panel",      cbClass:"phase-cb",      stateKey:"phases",     allLabel:"All Phases",     btnLabelId:"phase-btn-label",      selectAllId:"phase-select-all",      clearAllId:"phase-clear-all" }});
const statusDD    = makeMultiDD({{ btnId:"status-btn",     panelId:"status-panel",     cbClass:"status-cb",     stateKey:"statuses",   allLabel:"All Statuses",   btnLabelId:"status-btn-label",     selectAllId:"status-select-all",     clearAllId:"status-clear-all" }});
const countryDD   = makeMultiDD({{ btnId:"country-btn",    panelId:"country-panel",    cbClass:"country-cb",    stateKey:"countries",  allLabel:"All Countries",  btnLabelId:"country-btn-label",    selectAllId:"country-select-all",    clearAllId:"country-clear-all",  searchId:"country-search" }});
const intvTypeDD  = makeMultiDD({{ btnId:"intv-type-btn", panelId:"intv-type-panel", cbClass:"intv-type-cb", stateKey:"intTypes", allLabel:"All Types", btnLabelId:"intv-type-btn-label", selectAllId:"intv-type-select-all", clearAllId:"intv-type-clear-all" }});
const funderDD    = makeMultiDD({{ btnId:"funder-btn",     panelId:"funder-panel",     cbClass:"funder-cb",     stateKey:"funders",    allLabel:"All Funders",    btnLabelId:"funder-btn-label",     selectAllId:"funder-select-all",     clearAllId:"funder-clear-all" }});
const studyTypeDD = makeMultiDD({{ btnId:"study-type-btn", panelId:"study-type-panel", cbClass:"study-type-cb", stateKey:"studyTypes", allLabel:"All Types",      btnLabelId:"study-type-btn-label", selectAllId:"study-type-select-all", clearAllId:"study-type-clear-all" }});

// Close all panels when clicking outside
document.addEventListener("click", () => document.querySelectorAll(".multi-panel.open").forEach(p => p.classList.remove("open")));

// ── SEARCH ────────────────────────────────────────────────────────────────────
let searchTimer;
document.getElementById("search-input").addEventListener("input", e => {{
  clearTimeout(searchTimer);
  searchTimer = setTimeout(() => {{ state.search = e.target.value; applyFilters(); }}, 200);
}});

// ── DATE RANGE ────────────────────────────────────────────────────────────────
document.getElementById("date-from").addEventListener("change", e => {{ state.dateFrom = e.target.value; applyFilters(); }});
document.getElementById("date-to").addEventListener("change",   e => {{ state.dateTo   = e.target.value; applyFilters(); }});

// ── CLEAR ALL ─────────────────────────────────────────────────────────────────
document.getElementById("btn-clear").addEventListener("click", () => {{
  document.getElementById("search-input").value = "";
  document.getElementById("date-from").value    = "";
  document.getElementById("date-to").value      = "";
  state.search = ""; state.dateFrom = ""; state.dateTo = "";
  areaDD.reset(); phaseDD.reset(); statusDD.reset(); countryDD.reset(); intvTypeDD.reset(); funderDD.reset(); studyTypeDD.reset();
  applyFilters();
}});

// ── CSV EXPORT ────────────────────────────────────────────────────────────────
document.getElementById("btn-export").addEventListener("click", () => {{
  const cols = ["nct_id","title","sponsor","funder_types","study_type","therapy_areas","phase","status",
                "countries","intervention_types","interventions","conditions","start_date","completion_date","enrollment",
                "allocation","intervention_model","primary_purpose","masking","url"];
  const header = ["NCT ID","Title","Sponsor","Funder Type","Study Type","Therapy Areas","Phase","Status",
                  "Countries","Intervention Types","Interventions","Conditions","Start Date","Completion Date","Enrollment",
                  "Allocation","Intervention Model","Primary Purpose","Masking","URL"];

  const rows = [header, ...filtered.map(r =>
    cols.map(c => {{
      let v = r[c];
      if (Array.isArray(v)) v = v.join("; ");
      v = String(v ?? "").replace(/"/g,'""');
      return `"${{v}}"`;
    }})
  )];

  const blob = new Blob([rows.map(r=>r.join(",")).join("\\n")], {{type:"text/csv"}});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = `clinical_trials_${{new Date().toISOString().slice(0,10)}}.csv`;
  a.click();
}});

// ── XLSX EXPORT ───────────────────────────────────────────────────────────────
document.getElementById("btn-export-xlsx").addEventListener("click", () => {{
  const cols = ["nct_id","title","sponsor","funder_types","study_type","therapy_areas","phase","status",
                "countries","intervention_types","interventions","conditions","start_date","completion_date","enrollment",
                "allocation","intervention_model","primary_purpose","masking","url"];
  const header = ["NCT ID","Title","Sponsor","Funder Type","Study Type","Therapy Areas","Phase","Status",
                  "Countries","Intervention Types","Interventions","Conditions","Start Date","Completion Date","Enrollment",
                  "Allocation","Intervention Model","Primary Purpose","Masking","URL"];
  const rows = [header, ...filtered.map(r =>
    cols.map(c => Array.isArray(r[c]) ? r[c].join("; ") : (r[c] ?? ""))
  )];
  const ws = XLSX.utils.aoa_to_sheet(rows);
  // Column widths
  ws['!cols'] = [8,40,25,18,18,28,10,20,22,18,30,28,12,12,10,14,20,18,14,35].map(w => ({{wch:w}}));
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, "Trials");
  XLSX.writeFile(wb, `clinical_trials_${{new Date().toISOString().slice(0,10)}}.xlsx`);
}});

// ── SPONSORS EXPORT (deduplicated, SeamlessAI-ready) ─────────────────────────
function buildSponsorMap(source) {{
  const map = {{}};
  function ensure(name, org, funderTypes) {{
    if (!map[name]) map[name] = {{
      name, org_type: org||"", funder_types: new Set(funderTypes||[]),
      areas: new Set(), countries: new Set(),
      lead_trials: 0, collab_trials: 0,
      phases: new Set(), statuses: new Set(),
    }};
  }}
  source.forEach(r => {{
    if (r.sponsor) {{
      ensure(r.sponsor, r.org_type, r.funder_types);
      const e = map[r.sponsor];
      e.lead_trials++;
      (r.therapy_areas||[]).forEach(a => e.areas.add(a));
      (r.countries    ||[]).forEach(c => e.countries.add(c));
      (r.funder_types ||[]).forEach(f => e.funder_types.add(f));
      if (r.phase && r.phase !== "N/A") e.phases.add(r.phase);
      if (r.status) e.statuses.add(r.status);
    }}
    (r.collaborators||[]).forEach(c => {{
      if (!c) return;
      ensure(c, "", []);
      map[c].collab_trials++;
      (r.therapy_areas||[]).forEach(a => map[c].areas.add(a));
      (r.countries    ||[]).forEach(ct=> map[c].countries.add(ct));
      if (r.phase && r.phase !== "N/A") map[c].phases.add(r.phase);
      if (r.status) map[c].statuses.add(r.status);
    }});
  }});
  return Object.values(map)
    .map(e => ({{...e, total_trials: e.lead_trials + e.collab_trials}}))
    .sort((a,b) => b.total_trials - a.total_trials);
}}

const FUNDER_MAP = {{INDUSTRY:"Industry",NIH:"NIH",FED:"U.S. Fed Gov't (non-NIH)",OTHER_GOV:"Other Gov't",INDIV:"Individual",NETWORK:"Network",OTHER:"Other",UNKNOWN:"Unknown"}};

document.getElementById("btn-export-sponsors").addEventListener("click", () => {{
  const sponsors = buildSponsorMap(filtered);
  const header   = ["Company Name","Org Type","Funder Types","Role","Lead Trials","Collab Trials","Total Trials","Therapy Areas","Phases","Statuses","Countries"];
  const rows = sponsors.map(s => {{
    const orgLabel    = FUNDER_MAP[s.org_type] || s.org_type || "Other";
    const funderLabel = [...s.funder_types].map(f => FUNDER_MAP[f]||f).join("; ");
    const role        = s.lead_trials > 0 && s.collab_trials > 0 ? "Lead + Collaborator"
                      : s.lead_trials > 0 ? "Lead Sponsor" : "Collaborator";
    const areas       = [...s.areas].sort().join("; ");
    const countries   = [...s.countries].sort().join("; ");
    const phases      = [...s.phases].join("; ");
    const statuses    = [...s.statuses].join("; ");
    return [s.name, orgLabel, funderLabel, role, s.lead_trials, s.collab_trials, s.total_trials, areas, phases, statuses, countries];
  }});

  const ws = XLSX.utils.aoa_to_sheet([header, ...rows]);
  ws['!cols'] = [36,14,22,18,10,10,10,40,20,22,30].map(w => ({{wch:w}}));

  // Freeze header row
  ws['!freeze'] = {{xSplit:0, ySplit:1}};

  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, "Sponsors");
  const date = new Date().toISOString().slice(0,10);
  XLSX.writeFile(wb, `sponsors_seamlessai_${{date}}.xlsx`);
}});

function updateSponsorsBadge() {{
  const n = new Set(filtered.filter(r => r.sponsor).map(r => r.sponsor)).size;
  document.getElementById("sponsors-export-badge").textContent = n.toLocaleString();
}}

// ── TAB SWITCHING ─────────────────────────────────────────────────────────────
function switchTab(tab) {{
  document.getElementById("panel-trials").style.display    = tab === "trials"    ? "" : "none";
  document.getElementById("panel-sponsors").style.display  = tab === "sponsors"  ? "" : "none";
  document.getElementById("panel-analytics").style.display = tab === "analytics" ? "" : "none";
  document.getElementById("tab-trials").classList.toggle("active",    tab === "trials");
  document.getElementById("tab-sponsors").classList.toggle("active",  tab === "sponsors");
  document.getElementById("tab-analytics").classList.toggle("active", tab === "analytics");
  if (tab === "sponsors")  renderSponsors();
  if (tab === "analytics") renderAnalytics();
}}

// ── ANALYTICS ─────────────────────────────────────────────────────────────────
const _chartInstances = {{}};

function _destroyChart(id) {{
  if (_chartInstances[id]) {{ _chartInstances[id].destroy(); delete _chartInstances[id]; }}
}}

function _countBy(arr, keyFn) {{
  const m = {{}};
  arr.forEach(r => {{
    const keys = keyFn(r);
    (Array.isArray(keys) ? keys : [keys]).forEach(k => {{
      if (k) m[k] = (m[k] || 0) + 1;
    }});
  }});
  return m;
}}

function _topN(map, n) {{
  return Object.entries(map).sort((a,b) => b[1]-a[1]).slice(0,n);
}}

const _CHART_DEFAULTS = {{
  responsive: true,
  maintainAspectRatio: false,
  animation: {{ duration: 400 }},
  plugins: {{
    legend: {{ labels: {{ font: {{ size: 12, family: "'Inter', system-ui, sans-serif" }}, boxWidth: 12, padding: 14 }} }},
    tooltip: {{ bodyFont: {{ size: 12 }}, titleFont: {{ size: 12 }} }},
  }},
}};

const _PALETTE = [
  "#4f46e5","#0ea5e9","#10b981","#f59e0b","#ef4444","#8b5cf6",
  "#06b6d4","#84cc16","#f97316","#ec4899","#14b8a6","#a855f7",
  "#6366f1","#22d3ee","#facc15","#fb923c","#e879f9","#34d399",
];

const _FUNDER_LABELS_A = {{
  INDUSTRY:"Industry", NIH:"NIH", FED:"U.S. Fed (non-NIH)",
  OTHER_GOV:"Other Gov't", INDIV:"Individual", NETWORK:"Network",
  OTHER:"Other", UNKNOWN:"Unknown",
}};
const _INTV_TYPE_LABELS_A = {{
  DRUG:"Drug", BIOLOGICAL:"Biological", DEVICE:"Device",
  PROCEDURE:"Procedure", BEHAVIORAL:"Behavioral", RADIATION:"Radiation",
  DIETARY_SUPPLEMENT:"Dietary Supplement", GENETIC:"Genetic",
  COMBINATION_PRODUCT:"Combination Product", DIAGNOSTIC_TEST:"Diagnostic Test",
  OTHER:"Other",
}};
const _INTV_TYPE_COLORS_A = {{
  DRUG:"#dbeafe", BIOLOGICAL:"#ede9fe", DEVICE:"#d1fae5",
  PROCEDURE:"#fef3c7", BEHAVIORAL:"#e0f2fe", RADIATION:"#fee2e2",
  DIETARY_SUPPLEMENT:"#ecfccb", GENETIC:"#eef2ff",
  COMBINATION_PRODUCT:"#fce7f3", DIAGNOSTIC_TEST:"#f1f5f9", OTHER:"#f3f4f6",
}};
const _INTV_TYPE_BORDER_A = {{
  DRUG:"#1d4ed8", BIOLOGICAL:"#6d28d9", DEVICE:"#065f46",
  PROCEDURE:"#92400e", BEHAVIORAL:"#0369a1", RADIATION:"#b91c1c",
  DIETARY_SUPPLEMENT:"#3f6212", GENETIC:"#3730a3",
  COMBINATION_PRODUCT:"#9d174d", DIAGNOSTIC_TEST:"#475569", OTHER:"#6b7280",
}};

function renderAnalytics() {{
  const src = filtered;
  if (!src.length) return;

  // ── Summary stats ──────────────────────────────────────────────────────────
  const totalEnroll = src.reduce((s,r) => s + (parseInt(r.enrollment)||0), 0);
  document.getElementById("an-total").textContent      = src.length.toLocaleString();
  document.getElementById("an-recruiting").textContent = src.filter(r=>r.status==="RECRUITING").length.toLocaleString();
  document.getElementById("an-sponsors").textContent   = new Set(src.map(r=>r.sponsor).filter(Boolean)).size.toLocaleString();
  document.getElementById("an-countries").textContent  = new Set(src.flatMap(r=>r.countries)).size.toLocaleString();
  document.getElementById("an-enrollment").textContent = totalEnroll > 0 ? totalEnroll.toLocaleString() : "–";

  // ── 1. Monthly trend ───────────────────────────────────────────────────────
  (function() {{
    const m = {{}};
    src.forEach(r => {{
      if (!r.start_date) return;
      const ym = r.start_date.slice(0,7);
      if (/^\d{{4}}-\d{{2}}$/.test(ym)) m[ym] = (m[ym]||0) + 1;
    }});
    const labels = Object.keys(m).sort();
    const values = labels.map(l => m[l]);
    const pretty = labels.map(l => {{
      const [y,mo] = l.split("-");
      return new Date(+y, +mo-1).toLocaleString("default",{{month:"short",year:"2-digit"}});
    }});
    _destroyChart("trend");
    _chartInstances["trend"] = new Chart(document.getElementById("ch-trend"), {{
      type: "bar",
      data: {{
        labels: pretty,
        datasets: [{{ label:"Trials Started", data: values,
          backgroundColor: "rgba(79,70,229,.75)", borderColor:"#4f46e5",
          borderWidth:1, borderRadius:3 }}]
      }},
      options: {{ ..._CHART_DEFAULTS,
        plugins: {{ ..._CHART_DEFAULTS.plugins, legend:{{display:false}} }},
        scales: {{
          x: {{ grid:{{display:false}}, ticks:{{font:{{size:11}},maxRotation:45}} }},
          y: {{ grid:{{color:"#f1f5f9"}}, ticks:{{font:{{size:11}},precision:0}} }},
        }},
      }},
    }});
  }})();

  // ── 2. Therapy area horizontal bar ────────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.therapy_areas);
    const top = _topN(m, 15).reverse();
    const labels = top.map(([k]) => k);
    const values = top.map(([,v]) => v);
    const colors = labels.map(l => (AREA_COLORS[l] || "#78909c") + "cc");
    _destroyChart("area");
    _chartInstances["area"] = new Chart(document.getElementById("ch-area"), {{
      type: "bar",
      data: {{ labels, datasets:[{{ label:"Trials", data:values,
        backgroundColor:colors, borderColor:colors.map(c=>c.slice(0,7)),
        borderWidth:1, borderRadius:3 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        indexAxis:"y",
        plugins:{{ ..._CHART_DEFAULTS.plugins, legend:{{display:false}} }},
        scales:{{
          x:{{ grid:{{color:"#f1f5f9"}}, ticks:{{font:{{size:11}},precision:0}} }},
          y:{{ grid:{{display:false}}, ticks:{{font:{{size:11}}}} }},
        }},
      }},
    }});
  }})();

  // ── 3. Phase doughnut ─────────────────────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.phase || "N/A");
    const phaseOrder = ["Phase 1","Phase 1/Phase 2","Phase 2","Phase 2/Phase 3","Phase 3","Phase 3/Phase 4","Phase 4","N/A","Early Phase 1"];
    const entries = Object.entries(m).sort((a,b) => {{
      const ia = phaseOrder.indexOf(a[0]), ib = phaseOrder.indexOf(b[0]);
      return (ia<0?99:ia) - (ib<0?99:ib);
    }});
    const labels = entries.map(([k])=>k);
    const values = entries.map(([,v])=>v);
    const phaseColors = ["#42a5f5","#64b5f6","#1976d2","#1565c0","#0d47a1","#0a3470","#082a5e","#90a4ae","#90caf9"];
    _destroyChart("phase");
    _chartInstances["phase"] = new Chart(document.getElementById("ch-phase"), {{
      type: "doughnut",
      data: {{ labels, datasets:[{{ data:values,
        backgroundColor: labels.map((_,i) => phaseColors[i % phaseColors.length]),
        borderWidth:2, borderColor:"#fff" }}] }},
      options: {{ ..._CHART_DEFAULTS,
        cutout:"55%",
        plugins:{{ ..._CHART_DEFAULTS.plugins,
          legend:{{ position:"right", labels:{{font:{{size:11}},boxWidth:12,padding:10}} }} }},
      }},
    }});
  }})();

  // ── 4. Status doughnut ────────────────────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.status);
    const entries = _topN(m, 10);
    const labels = entries.map(([k]) => k.replace(/_/g," ").replace(/\b\w/g,x=>x.toUpperCase()));
    const values = entries.map(([,v]) => v);
    const rawKeys = entries.map(([k]) => k);
    const colors  = rawKeys.map(k => STATUS_COLORS[k] || "#94a3b8");
    _destroyChart("status");
    _chartInstances["status"] = new Chart(document.getElementById("ch-status"), {{
      type: "doughnut",
      data: {{ labels, datasets:[{{ data:values,
        backgroundColor: colors.map(c => c+"cc"),
        borderColor: colors, borderWidth:2 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        cutout:"52%",
        plugins:{{ ..._CHART_DEFAULTS.plugins,
          legend:{{ position:"right", labels:{{font:{{size:11}},boxWidth:12,padding:8}} }} }},
      }},
    }});
  }})();

  // ── 5. Funder type doughnut ───────────────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.funder_types||[]);
    const entries = _topN(m, 8);
    const labels = entries.map(([k]) => _FUNDER_LABELS_A[k] || k);
    const values = entries.map(([,v]) => v);
    const fColors = ["#4f46e5","#0ea5e9","#10b981","#f59e0b","#ef4444","#8b5cf6","#06b6d4","#94a3b8"];
    _destroyChart("funder");
    _chartInstances["funder"] = new Chart(document.getElementById("ch-funder"), {{
      type: "doughnut",
      data: {{ labels, datasets:[{{ data:values,
        backgroundColor: fColors.map(c=>c+"cc"),
        borderColor: fColors, borderWidth:2 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        cutout:"52%",
        plugins:{{ ..._CHART_DEFAULTS.plugins,
          legend:{{ position:"right", labels:{{font:{{size:11}},boxWidth:12,padding:8}} }} }},
      }},
    }});
  }})();

  // ── 6. Intervention type horizontal bar ───────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.intervention_types||[]);
    const order = ["DRUG","BIOLOGICAL","DEVICE","PROCEDURE","BEHAVIORAL","RADIATION","DIETARY_SUPPLEMENT","GENETIC","COMBINATION_PRODUCT","DIAGNOSTIC_TEST","OTHER"];
    const entries = Object.entries(m).sort((a,b) => {{
      const ia=order.indexOf(a[0]), ib=order.indexOf(b[0]);
      return (ia<0?99:ia)-(ib<0?99:ib);
    }}).reverse();
    const labels = entries.map(([k]) => _INTV_TYPE_LABELS_A[k]||k);
    const values = entries.map(([,v]) => v);
    const rawKeys = entries.map(([k]) => k);
    _destroyChart("intv-type");
    _chartInstances["intv-type"] = new Chart(document.getElementById("ch-intv-type"), {{
      type: "bar",
      data: {{ labels, datasets:[{{ label:"Trials", data:values,
        backgroundColor: rawKeys.map(k=>_INTV_TYPE_COLORS_A[k]||"#f1f5f9"),
        borderColor:     rawKeys.map(k=>_INTV_TYPE_BORDER_A[k]||"#94a3b8"),
        borderWidth:1, borderRadius:4 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        indexAxis:"y",
        plugins:{{ ..._CHART_DEFAULTS.plugins, legend:{{display:false}} }},
        scales:{{
          x:{{ grid:{{color:"#f1f5f9"}}, ticks:{{font:{{size:11}},precision:0}} }},
          y:{{ grid:{{display:false}}, ticks:{{font:{{size:11}}}} }},
        }},
      }},
    }});
  }})();

  // ── 7. Study type doughnut ────────────────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.study_type ? r.study_type.replace(/_/g," ") : "Unknown");
    const entries = _topN(m, 8);
    const labels = entries.map(([k])=>k);
    const values = entries.map(([,v])=>v);
    const stColors = ["#4f46e5","#10b981","#f59e0b","#ef4444","#8b5cf6","#0ea5e9","#94a3b8","#06b6d4"];
    _destroyChart("study-type");
    _chartInstances["study-type"] = new Chart(document.getElementById("ch-study-type"), {{
      type: "doughnut",
      data: {{ labels, datasets:[{{ data:values,
        backgroundColor: stColors.map(c=>c+"cc"),
        borderColor: stColors, borderWidth:2 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        cutout:"52%",
        plugins:{{ ..._CHART_DEFAULTS.plugins,
          legend:{{ position:"right", labels:{{font:{{size:11}},boxWidth:12,padding:8}} }} }},
      }},
    }});
  }})();

  // ── 8. Top countries horizontal bar ───────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.countries||[]);
    const top = _topN(m, 15).reverse();
    const labels = top.map(([k])=>k);
    const values = top.map(([,v])=>v);
    _destroyChart("country");
    _chartInstances["country"] = new Chart(document.getElementById("ch-country"), {{
      type: "bar",
      data: {{ labels, datasets:[{{ label:"Trials", data:values,
        backgroundColor:"rgba(14,165,233,.75)", borderColor:"#0ea5e9",
        borderWidth:1, borderRadius:3 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        indexAxis:"y",
        plugins:{{ ..._CHART_DEFAULTS.plugins, legend:{{display:false}} }},
        scales:{{
          x:{{ grid:{{color:"#f1f5f9"}}, ticks:{{font:{{size:11}},precision:0}} }},
          y:{{ grid:{{display:false}}, ticks:{{font:{{size:11}}}} }},
        }},
      }},
    }});
  }})();

  // ── 9. Top sponsors horizontal bar ────────────────────────────────────────
  (function() {{
    const m = _countBy(src, r => r.sponsor);
    const top = _topN(m, 15).reverse();
    const labels = top.map(([k])=>k.length>30?k.slice(0,30)+"…":k);
    const values = top.map(([,v])=>v);
    _destroyChart("sponsor");
    _chartInstances["sponsor"] = new Chart(document.getElementById("ch-sponsor"), {{
      type: "bar",
      data: {{ labels, datasets:[{{ label:"Trials", data:values,
        backgroundColor:"rgba(16,185,129,.75)", borderColor:"#10b981",
        borderWidth:1, borderRadius:3 }}] }},
      options: {{ ..._CHART_DEFAULTS,
        indexAxis:"y",
        plugins:{{ ..._CHART_DEFAULTS.plugins, legend:{{display:false}} }},
        scales:{{
          x:{{ grid:{{color:"#f1f5f9"}}, ticks:{{font:{{size:11}},precision:0}} }},
          y:{{ grid:{{display:false}}, ticks:{{font:{{size:11}}}} }},
        }},
      }},
    }});
  }})();
}}

// ── SPONSORS PANEL ────────────────────────────────────────────────────────────
const spAreas = new Set();
const spAreaBtn   = document.getElementById("sp-area-btn");
const spAreaPanel = document.getElementById("sp-area-panel");
spAreaBtn.addEventListener("click", e => {{
  e.stopPropagation();
  document.querySelectorAll(".multi-panel.open").forEach(p => p !== spAreaPanel && p.classList.remove("open"));
  spAreaPanel.classList.toggle("open");
}});
spAreaPanel.addEventListener("click", e => e.stopPropagation());
document.getElementById("sp-area-all").addEventListener("click", () => {{
  spAreaPanel.querySelectorAll(".sp-area-cb").forEach(cb => cb.checked = true);
  spAreas.clear();
  [...spAreaPanel.querySelectorAll(".sp-area-cb")].forEach(cb => spAreas.add(cb.value));
  updateSpAreaLabel(); renderSponsors();
}});
document.getElementById("sp-area-clear").addEventListener("click", () => {{
  spAreaPanel.querySelectorAll(".sp-area-cb").forEach(cb => cb.checked = false);
  spAreas.clear(); updateSpAreaLabel(); renderSponsors();
}});
spAreaPanel.querySelectorAll(".sp-area-cb").forEach(cb => {{
  cb.addEventListener("change", () => {{
    if (cb.checked) spAreas.add(cb.value); else spAreas.delete(cb.value);
    updateSpAreaLabel(); renderSponsors();
  }});
}});
function updateSpAreaLabel() {{
  const n = spAreas.size;
  document.getElementById("sp-area-label").textContent = n === 0 ? "All Areas" : n === 1 ? [...spAreas][0] : `${{n}} areas`;
  spAreaBtn.classList.toggle("active", n > 0);
}}

function renderSponsors() {{
  const q   = (document.getElementById("sponsor-search")?.value || "").toLowerCase();
  const src = spAreas.size > 0
    ? ALL_DATA.filter(r => r.therapy_areas.some(a => spAreas.has(a)))
    : ALL_DATA;

  // Build unified company map — track lead sponsorships AND collaborations separately
  const map = {{}};
  function ensureEntry(name, org) {{
    if (!map[name]) map[name] = {{ lead:0, collab:0, areas:new Set(), countries:new Set(), org: org||"OTHER" }};
  }}
  src.forEach(r => {{
    if (r.sponsor) {{
      ensureEntry(r.sponsor, r.org_type);
      map[r.sponsor].lead++;
      r.therapy_areas.forEach(a => map[r.sponsor].areas.add(a));
      r.countries.forEach(c => map[r.sponsor].countries.add(c));
    }}
    (r.collaborators||[]).forEach(c => {{
      if (!c) return;
      ensureEntry(c, "");
      map[c].collab++;
      r.therapy_areas.forEach(a => map[c].areas.add(a));
      r.countries.forEach(ct => map[c].countries.add(ct));
    }});
  }});

  let sorted = Object.entries(map)
    .map(([name, info]) => [name, {{...info, total: info.lead + info.collab}}])
    .sort((a,b) => b[1].total - a[1].total);
  if (q) sorted = sorted.filter(([name]) => name.toLowerCase().includes(q));

  document.getElementById("sponsors-showing").textContent = sorted.length.toLocaleString();
  document.getElementById("tab-sponsors-badge").textContent = sorted.length.toLocaleString();

  const grid = document.getElementById("sponsors-grid");
  if (!sorted.length) {{
    grid.innerHTML = `<div style="grid-column:1/-1;text-align:center;padding:60px;color:var(--text-muted)">No companies match the current filter.</div>`;
    return;
  }}

  grid.innerHTML = sorted.map(([name, info]) => {{
    const areas    = [...info.areas].slice(0,3).map(areaTag).join("");
    const orgColor = info.org === "INDUSTRY" ? "#4f46e5" : info.org === "NIH" ? "#059669" : "#78909c";
    // Role badge: Lead / Collab / Both
    const roleBadge = info.lead > 0 && info.collab > 0
      ? `<span class="role-badge role-both">Lead + Collab</span>`
      : info.lead > 0
        ? `<span class="role-badge role-lead">Lead Sponsor</span>`
        : `<span class="role-badge role-collab">Collaborator</span>`;
    const metaParts = [];
    if (info.lead  > 0) metaParts.push(`${{info.lead}} lead`);
    if (info.collab> 0) metaParts.push(`${{info.collab}} collab`);
    return `<div class="sponsor-card" onclick="filterBySponsor('${{name.replace(/'/g,"\\'")}}')">
      <div class="sponsor-trial-count">${{info.total}}</div>
      <div style="margin-bottom:3px">${{roleBadge}}</div>
      <div style="display:flex;align-items:center;gap:6px;margin-bottom:4px">
        <span style="width:7px;height:7px;border-radius:50%;background:${{orgColor}};flex-shrink:0"></span>
        <div class="sponsor-card-name">${{name}}</div>
      </div>
      <div class="sponsor-card-meta">${{metaParts.join(" · ")}} &nbsp;·&nbsp; ${{info.countries.size}} countr${{info.countries.size>1?"ies":"y"}}</div>
      <div class="sponsor-card-areas">${{areas}}</div>
    </div>`;
  }}).join("");
}}

function filterBySponsor(name) {{
  // Switch to trials tab and set search to this sponsor
  switchTab("trials");
  const si = document.getElementById("search-input");
  si.value = name;
  state.search = name;
  applyFilters();
  si.focus();
}}

// ── KPI: RECRUITING COUNT ─────────────────────────────────────────────────────
function updateRecruitingKPI() {{
  const n = filtered.filter(r => r.status === "RECRUITING").length;
  document.getElementById("kpi-recruiting").textContent = n.toLocaleString();
  document.getElementById("kpi-recruiting-sub").textContent = "Actively recruiting";
}}

// ── INITIAL RENDER ────────────────────────────────────────────────────────────
// Default sort: newest start date first
document.querySelector('th[data-col="start_date"]').classList.add("sort-desc");
applyFilters();
updateRecruitingKPI();

</script>

<!-- FILE LOAD OVERLAY (viewer mode only) -->
<div id="load-overlay" style="display:{'none' if not viewer_mode else 'flex'};position:fixed;inset:0;z-index:9999;background:#f0f2f5;flex-direction:column;align-items:center;justify-content:center;font-family:'Inter',system-ui,sans-serif">
  <div style="background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:48px 52px;max-width:520px;width:90%;text-align:center;box-shadow:0 8px 32px rgba(0,0,0,.1)">
    <div style="font-size:40px;margin-bottom:16px">📊</div>
    <div style="font-size:22px;font-weight:700;color:#0f172a;margin-bottom:8px">Clinical Trials Dashboard</div>
    <div style="font-size:14px;color:#718096;margin-bottom:28px;line-height:1.6">Drop your <code style="background:#f1f5f9;padding:2px 6px;border-radius:4px;font-size:13px">trials_data.json</code> file here,<br>or click to browse</div>
    <div id="drop-zone" style="border:2px dashed #c7d2fe;border-radius:10px;padding:32px 20px;cursor:pointer;transition:all .2s;background:#f8faff"
         ondragover="event.preventDefault();this.style.borderColor='#4f46e5';this.style.background='#eef2ff'"
         ondragleave="this.style.borderColor='#c7d2fe';this.style.background='#f8faff'"
         ondrop="event.preventDefault();this.style.borderColor='#c7d2fe';this.style.background='#f8faff';handleFile(event.dataTransfer.files[0])">
      <div style="font-size:32px;margin-bottom:8px">⬆</div>
      <div style="font-size:14px;font-weight:600;color:#4f46e5">Click to select file</div>
      <div style="font-size:12px;color:#94a3b8;margin-top:4px">or drag &amp; drop</div>
      <input type="file" id="file-input" accept=".json" style="display:none" onchange="handleFile(this.files[0])">
    </div>
    <div id="load-error" style="display:none;margin-top:14px;padding:10px 14px;background:#fee2e2;border-radius:8px;color:#b91c1c;font-size:13px"></div>
    <div style="margin-top:20px;font-size:11px;color:#94a3b8">Generated by fetch_trials.py &nbsp;·&nbsp; clinicaltrials.gov API v2</div>
  </div>
</div>

<script>
// ── VIEWER MODE: FILE LOADER + BOOTSTRAP ─────────────────────────────────────
document.getElementById("drop-zone").addEventListener("click", () => document.getElementById("file-input").click());

function handleFile(file) {{
  if (!file) return;
  if (!file.name.endsWith(".json")) {{
    document.getElementById("load-error").style.display = "";
    document.getElementById("load-error").textContent   = "Please select a .json file generated by fetch_trials.py";
    return;
  }}
  const reader = new FileReader();
  reader.onload = ev => {{
    try {{
      const data = JSON.parse(ev.target.result);
      bootstrapDashboard(data);
    }} catch(e) {{
      document.getElementById("load-error").style.display = "";
      document.getElementById("load-error").textContent   = "Could not parse file: " + e.message;
    }}
  }};
  reader.readAsText(file);
}}

const _FUNDER_LABELS_V = {{INDUSTRY:"Industry",NIH:"NIH",FED:"U.S. Fed Gov't (non-NIH)",OTHER_GOV:"Other Gov't",INDIV:"Individual",NETWORK:"Network",OTHER:"Other",UNKNOWN:"Unknown"}};
const _INTV_LABELS_V   = {{DRUG:"Drug",BIOLOGICAL:"Biological",DEVICE:"Device",PROCEDURE:"Procedure",BEHAVIORAL:"Behavioral",RADIATION:"Radiation",DIETARY_SUPPLEMENT:"Dietary Supplement",GENETIC:"Genetic",COMBINATION_PRODUCT:"Combination Product",DIAGNOSTIC_TEST:"Diagnostic Test",OTHER:"Other"}};

function buildFilterPanels(records) {{
  const uniq = arr => [...new Set(arr.filter(Boolean))].sort();
  function cbs(items, cls, labelFn) {{
    return items.map(v => `<label class="cb-label"><input type="checkbox" class="${{cls}}" value="${{v}}"> ${{labelFn ? labelFn(v) : v}}</label>`).join("\\n");
  }}
  function inject(panelId, cbHtml, selId, clrId, searchId) {{
    const search = searchId ? `<input type="text" id="${{searchId}}" placeholder="Search…" style="width:100%;border:1px solid var(--border);border-radius:5px;padding:5px 8px;font-size:12px;margin-bottom:6px;box-sizing:border-box">` : "";
    document.getElementById(panelId).innerHTML = search + cbHtml + `<div class="panel-actions"><a id="${{selId}}">All</a><a id="${{clrId}}">Clear</a></div>`;
  }}

  const areas     = uniq(records.flatMap(r => r.therapy_areas   || []));
  const phases    = uniq(records.map(r => r.phase).filter(p => p && p !== "N/A"));
  const statuses  = uniq(records.map(r => r.status));
  const countries = uniq(records.flatMap(r => r.countries       || []));
  const funders   = uniq(records.flatMap(r => r.funder_types    || []));
  const studyTypes= uniq(records.map(r => r.study_type));
  const intvTypes = uniq(records.flatMap(r => r.intervention_types || []));

  inject("area-panel",       cbs(areas,      "area-cb",       null),                                                           "area-select-all",       "area-clear-all");
  inject("phase-panel",      cbs(phases,     "phase-cb",      null),                                                           "phase-select-all",      "phase-clear-all");
  inject("status-panel",     cbs(statuses,   "status-cb",     s=>s.replace(/_/g," ").replace(/\b\w/g,x=>x.toUpperCase())),    "status-select-all",     "status-clear-all");
  inject("country-panel",    cbs(countries,  "country-cb",    null),                                                           "country-select-all",    "country-clear-all", "country-search");
  inject("funder-panel",     cbs(funders,    "funder-cb",     f=>_FUNDER_LABELS_V[f]||f),                                     "funder-select-all",     "funder-clear-all");
  inject("study-type-panel", cbs(studyTypes, "study-type-cb", s=>s.replace(/_/g," ").replace(/\b\w/g,x=>x.toUpperCase())),    "study-type-select-all", "study-type-clear-all");
  inject("intv-type-panel",  cbs(intvTypes,  "intv-type-cb",  t=>_INTV_LABELS_V[t]||t),                                       "intv-type-select-all",  "intv-type-clear-all");

  // Sponsors area filter
  const spPanel = document.getElementById("sp-area-panel");
  if (spPanel) spPanel.innerHTML = areas.map(a => `<label class="cb-label"><input type="checkbox" class="sp-area-cb" value="${{a}}"> ${{a}}</label>`).join("\\n") + `<div class="panel-actions"><a id="sp-area-all">All</a><a id="sp-area-clear">Clear</a></div>`;
}}

function rewireDropdowns() {{
  // Re-initialise all makeMultiDD instances after checkboxes are in the DOM
  window._areaDD     = makeMultiDD({{ btnId:"area-btn",       panelId:"area-panel",       cbClass:"area-cb",       stateKey:"areas",      allLabel:"All Areas",      btnLabelId:"area-btn-label",       selectAllId:"area-select-all",       clearAllId:"area-clear-all" }});
  window._phaseDD    = makeMultiDD({{ btnId:"phase-btn",      panelId:"phase-panel",      cbClass:"phase-cb",      stateKey:"phases",     allLabel:"All Phases",     btnLabelId:"phase-btn-label",      selectAllId:"phase-select-all",      clearAllId:"phase-clear-all" }});
  window._statusDD   = makeMultiDD({{ btnId:"status-btn",     panelId:"status-panel",     cbClass:"status-cb",     stateKey:"statuses",   allLabel:"All Statuses",   btnLabelId:"status-btn-label",     selectAllId:"status-select-all",     clearAllId:"status-clear-all" }});
  window._countryDD  = makeMultiDD({{ btnId:"country-btn",    panelId:"country-panel",    cbClass:"country-cb",    stateKey:"countries",  allLabel:"All Countries",  btnLabelId:"country-btn-label",    selectAllId:"country-select-all",    clearAllId:"country-clear-all", searchId:"country-search" }});
  window._intvTypeDD = makeMultiDD({{ btnId:"intv-type-btn",  panelId:"intv-type-panel",  cbClass:"intv-type-cb",  stateKey:"intTypes",   allLabel:"All Types",      btnLabelId:"intv-type-btn-label",  selectAllId:"intv-type-select-all",  clearAllId:"intv-type-clear-all" }});
  window._funderDD   = makeMultiDD({{ btnId:"funder-btn",     panelId:"funder-panel",     cbClass:"funder-cb",     stateKey:"funders",    allLabel:"All Funders",    btnLabelId:"funder-btn-label",     selectAllId:"funder-select-all",     clearAllId:"funder-clear-all" }});
  window._studyTypeDD= makeMultiDD({{ btnId:"study-type-btn", panelId:"study-type-panel", cbClass:"study-type-cb", stateKey:"studyTypes", allLabel:"All Types",      btnLabelId:"study-type-btn-label", selectAllId:"study-type-select-all", clearAllId:"study-type-clear-all" }});

  // Rewire Sponsors area DD
  const spBtn  = document.getElementById("sp-area-btn");
  const spPanl = document.getElementById("sp-area-panel");
  if (spBtn && spPanl) {{
    spBtn.addEventListener("click", e => {{
      e.stopPropagation();
      document.querySelectorAll(".multi-panel.open").forEach(p => p !== spPanl && p.classList.remove("open"));
      spPanl.classList.toggle("open");
    }});
    spPanl.addEventListener("click", e => e.stopPropagation());
    document.getElementById("sp-area-all")?.addEventListener("click",   () => {{ spPanl.querySelectorAll(".sp-area-cb").forEach(c=>c.checked=true);  spAreas.clear(); [...spPanl.querySelectorAll(".sp-area-cb")].forEach(c=>spAreas.add(c.value)); updateSpAreaLabel(); renderSponsors(); }});
    document.getElementById("sp-area-clear")?.addEventListener("click", () => {{ spPanl.querySelectorAll(".sp-area-cb").forEach(c=>c.checked=false); spAreas.clear(); updateSpAreaLabel(); renderSponsors(); }});
    spPanl.querySelectorAll(".sp-area-cb").forEach(cb => cb.addEventListener("change", () => {{ if(cb.checked) spAreas.add(cb.value); else spAreas.delete(cb.value); updateSpAreaLabel(); renderSponsors(); }}));
  }}

  // Rewire Clear-All button to use new DD references
  document.getElementById("btn-clear").replaceWith(document.getElementById("btn-clear").cloneNode(true));
  document.getElementById("btn-clear").addEventListener("click", () => {{
    document.getElementById("search-input").value = "";
    document.getElementById("date-from").value    = "";
    document.getElementById("date-to").value      = "";
    state.search = ""; state.dateFrom = ""; state.dateTo = "";
    [window._areaDD, window._phaseDD, window._statusDD, window._countryDD, window._intvTypeDD, window._funderDD, window._studyTypeDD].forEach(dd => dd && dd.reset());
    applyFilters();
  }});
}}

function bootstrapDashboard(data) {{
  // Populate global data
  ALL_DATA     = data.records || data;
  AREA_COLORS  = data.area_colors  || {{}};
  STATUS_COLORS= data.status_colors|| {{}};

  // Update header meta
  const m = data.meta || {{}};
  const metaEl = document.getElementById("meta-refresh");
  if (metaEl) metaEl.textContent = [
    m.generated_at ? "Refreshed: " + m.generated_at : "",
    (m.since_date && m.until_date) ? m.since_date + " – " + m.until_date : "",
    m.total ? m.total.toLocaleString() + " trials" : ""
  ].filter(Boolean).join("  ·  ");

  // Update date range label in KPI
  const kpiSub = document.querySelector(".kpi-sub");
  if (kpiSub && m.since_date && m.until_date) kpiSub.textContent = "Start date: " + m.since_date + " – " + m.until_date;

  // Add reload button to header
  const hdBtns = document.querySelector(".header-meta");
  if (hdBtns && !document.getElementById("btn-reload")) {{
    hdBtns.insertAdjacentHTML("beforeend",
      `<br><button id="btn-reload" onclick="document.getElementById('file-input-reload').click()" style="margin-top:4px;font-size:11px;border:1px solid #c7d2fe;background:#eef2ff;color:#4f46e5;border-radius:5px;padding:2px 8px;cursor:pointer;font-weight:600">⟳ Load different file</button>` +
      `<input type="file" id="file-input-reload" accept=".json" style="display:none" onchange="handleFile(this.files[0])">`
    );
  }}

  // Build filter panels, rewire dropdowns, hide overlay, render
  buildFilterPanels(ALL_DATA);
  rewireDropdowns();
  document.getElementById("load-overlay").style.display = "none";
  sortCol = "start_date"; sortDir = "desc";
  applyFilters();
  updateRecruitingKPI();
}}
</script>
""" + ("""
<script>
// AUTO-FETCH: when hosted on a web server (GitHub Pages), load data automatically
(function() {{
  fetch('./data/trials_data.json')
    .then(function(r) {{ if (!r.ok) throw new Error(r.status); return r.json(); }})
    .then(function(data) {{
      bootstrapDashboard(data);
    }})
    .catch(function() {{
      console.log('Auto-fetch unavailable — showing file picker');
    }});
}})();
</script>
""" if github_mode else "") + """
</body>
</html>"""
    return html


# ─── CSV / XLSX PROCESSING ─────────────────────────────────────────────────────
# Column names as exported by ClinicalTrials.gov "Download" button
CSV_COL_MAP = {
    # CSV header         → internal key
    "NCT Number":           "nct_id",
    "Title":                "title",
    "Status":               "status",
    "Study Results":        None,
    "Conditions":           "conditions_raw",
    "Interventions":        "interventions_raw",
    "Sponsor/Collaborators":"sponsor_raw",
    "Funded Bys":           "funded_by",
    "Phases":               "phase",
    "Enrollment":           "enrollment",
    "Study Type":           "study_type",
    "Start Date":           "start_date",
    "Primary Completion Date": "completion_date",
    "Last Update Posted":   "last_update",
    "Locations":            "locations_raw",
    "URL":                  "url",
}

# ─── JSON DATA EXPORT ──────────────────────────────────────────────────────────
def generate_json(records, generated_at, days=365, since_date=None, until_date=None):
    """Output processed trial data as a compact JSON file for the permanent viewer."""
    from datetime import timedelta
    _today = datetime.now()
    if since_date is None:
        since_date = (_today - timedelta(days=days)).strftime("%Y-%m-%d")
    if until_date is None:
        until_date = _today.strftime("%Y-%m-%d")
    payload = {
        "meta": {
            "generated_at":   generated_at,
            "since_date":     since_date,
            "until_date":     until_date,
            "days":           days,
            "total":          len(records),
            "schema_version": "1.0",
        },
        "area_colors":   AREA_COLORS,
        "status_colors": STATUS_COLORS,
        "records":       records,
    }
    return json.dumps(payload, separators=(",", ":"))

def parse_csv_file(path):
    """Parse a ClinicalTrials.gov CSV or XLSX export into the same record format."""
    path = Path(path)
    print(f"\n{'='*62}")
    print(f"  CSV / XLSX Mode  |  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  File: {path.name}")
    print(f"{'='*62}")

    rows = []
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                                   "--break-system-packages", "-q"])
            import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])))
        wb.close()
    else:
        # CSV (UTF-8 or latin-1)
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = path.read_text(encoding=enc)
                reader = csv.DictReader(io.StringIO(text))
                rows = [dict(r) for r in reader]
                break
            except UnicodeDecodeError:
                continue

    if not rows:
        print("  ✗ Could not read file.")
        return []

    print(f"  Loaded {len(rows):,} rows. Columns: {list(rows[0].keys())[:8]} …")

    # Filter: interventional drug/biologic, exclude observational/device
    keep_statuses = {s.upper() for s in TARGET_STATUSES} | {"RECRUITING", "ACTIVE, NOT RECRUITING", "NOT YET RECRUITING"}
    records = []

    for row in rows:
        # Normalise status
        status_raw = (row.get("Status") or row.get("status") or "").strip().upper()
        status_map = {
            "RECRUITING":                  "RECRUITING",
            "ACTIVE, NOT RECRUITING":      "ACTIVE_NOT_RECRUITING",
            "NOT YET RECRUITING":          "NOT_YET_RECRUITING",
            "ENROLLING BY INVITATION":     "ENROLLING_BY_INVITATION",
            "ACTIVE_NOT_RECRUITING":       "ACTIVE_NOT_RECRUITING",
            "NOT_YET_RECRUITING":          "NOT_YET_RECRUITING",
            "ENROLLING_BY_INVITATION":     "ENROLLING_BY_INVITATION",
        }
        status = status_map.get(status_raw, status_raw)
        if status not in {s.upper() for s in TARGET_STATUSES} and \
           status_raw not in {"RECRUITING","ACTIVE, NOT RECRUITING","NOT YET RECRUITING","ENROLLING BY INVITATION"}:
            continue

        study_type = (row.get("Study Type") or row.get("study_type") or "").strip()

        # Parse ALL interventions — extract both type and name
        intv_raw   = row.get("Interventions") or row.get("interventions_raw") or ""
        intv_parts = [p.strip() for p in intv_raw.split("|") if p.strip()]
        all_intvs  = []
        csv_intv_types = set()
        _TYPE_NORM = {
            "drug": "DRUG", "biological": "BIOLOGICAL", "device": "DEVICE",
            "procedure": "PROCEDURE", "behavioral": "BEHAVIORAL", "radiation": "RADIATION",
            "dietary supplement": "DIETARY_SUPPLEMENT", "genetic": "GENETIC",
            "combination product": "COMBINATION_PRODUCT", "diagnostic test": "DIAGNOSTIC_TEST",
            "other": "OTHER",
        }
        for part in intv_parts:
            m = re.match(r"^([\w][\w\s]+):\s*(.+)$", part)
            if m:
                t_raw = m.group(1).strip().lower()
                if t_raw in _TYPE_NORM:
                    csv_intv_types.add(_TYPE_NORM[t_raw])
                name = m.group(2).strip()
            else:
                name = part.strip()
            if name:
                all_intvs.append(name)

        # NCT ID + title
        nct_id = (row.get("NCT Number") or row.get("nct_id") or "").strip()
        title  = (row.get("Title")      or row.get("title")  or "").strip()
        url    = row.get("URL") or row.get("url") or (f"https://clinicaltrials.gov/study/{nct_id}" if nct_id else "")

        # Sponsor — ClinicalTrials.gov CSV puts "LeadSponsor|Collab1|Collab2"
        spons_raw = row.get("Sponsor/Collaborators") or row.get("sponsor_raw") or ""
        spons_parts = [s.strip() for s in spons_raw.split("|") if s.strip()]
        sponsor  = spons_parts[0] if spons_parts else ""
        collabs  = spons_parts[1:4]

        # Conditions
        cond_raw = row.get("Conditions") or row.get("conditions_raw") or ""
        conds    = [c.strip() for c in cond_raw.split("|") if c.strip()][:6]

        # Phase
        phase_raw = (row.get("Phases") or row.get("phase") or "N/A").strip()
        phase = phase_raw.replace("Phase ", "Phase ").strip() or "N/A"

        # Enrollment
        enrollment = str(row.get("Enrollment") or "").strip()

        # Dates
        start  = str(row.get("Start Date")               or "").strip()
        end    = str(row.get("Primary Completion Date")   or "").strip()
        updated= str(row.get("Last Update Posted")        or "").strip()

        # Countries — "City, State, Country|City2, Country2"
        locs_raw  = row.get("Locations") or row.get("locations_raw") or ""
        loc_parts = [l.strip() for l in locs_raw.split("|") if l.strip()]
        countries = sorted(set(
            lp.split(",")[-1].strip() for lp in loc_parts if "," in lp
        ))

        # Funded Bys — keep as list; also derive org_type from it
        funded_by_raw = row.get("Funded Bys") or row.get("funded_by") or ""
        funder_types  = [f.strip() for f in funded_by_raw.split("|") if f.strip()]
        funded_upper  = funded_by_raw.upper()
        if "INDUSTRY" in funded_upper:
            org_type = "INDUSTRY"
        elif "NIH" in funded_upper:
            org_type = "NIH"
        else:
            org_type = "OTHER"

        # Summary — not in standard CSV export
        summary = ""

        ta = classify_therapy_areas(conds, title, summary, interventions=all_intvs)

        records.append({
            "nct_id":               nct_id,
            "title":                title,
            "official_title":       "",
            "acronym":              "",
            "status":               status,
            "phase":                phase,
            "study_type":           study_type,
            "sponsor":              sponsor,
            "org_type":             org_type,
            "funder_types":         funder_types,
            "collaborators":        collabs,
            "conditions":           conds,
            "interventions":        all_intvs[:8],
            "intervention_types":   sorted(csv_intv_types),
            "countries":            countries,
            "therapy_areas":        ta,
            "start_date":           start,
            "completion_date":      end,
            "last_update":          updated,
            "enrollment":           enrollment,
            "summary":              summary,
            "url":                  url,
            "eligibility_criteria": "",
            "min_age":              "",
            "max_age":              "",
            "sex":                  "",
            "healthy_volunteers":   "",
            "primary_outcomes":     [],
            "secondary_outcomes":   [],
            "contacts":             [],
            "allocation":           "",
            "intervention_model":   "",
            "primary_purpose":      "",
            "masking":              "",
        })

    print(f"  ✓ Parsed {len(records):,} trials from CSV/XLSX.")
    return records


# ─── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="ClinicalTrials.gov Dashboard Generator")
    parser.add_argument("--max-pages", type=int, default=None,
                        help="Max API pages (1 page = 1000 records). Default: None = fetch all matching trials")
    parser.add_argument("--days", type=int, default=365,
                        help="Only fetch trials with a start date within this many days. Default: 365")
    parser.add_argument("--test", action="store_true",
                        help="Quick test: fetch only 1 API page (~1000 records)")
    parser.add_argument("--csv", type=str, default=None,
                        help="Path to a ClinicalTrials.gov CSV or XLSX export (skips API)")
    parser.add_argument("--output", type=str, default=str(OUTPUT_HTML),
                        help="Output HTML path")
    parser.add_argument("--json", type=str, default=None,
                        help="Output processed data as JSON (for use with permanent viewer). E.g. --json data/trials_data.json")
    parser.add_argument("--viewer", type=str, default=None,
                        help="Generate the permanent viewer HTML (run once). E.g. --viewer index.html")
    parser.add_argument("--use-llm", action="store_true", default=None,
                        help="Force LLM classification (requires ANTHROPIC_API_KEY env var)")
    parser.add_argument("--no-llm", action="store_true",
                        help="Force keyword-only classification mode")
    parser.add_argument("--cache-path", type=str, default="classification_cache.json",
                        help="Path to LLM classification cache file. Default: classification_cache.json")
    args = parser.parse_args()

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Determine LLM mode ────────────────────────────────────────────────────
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip() or None  # treat empty string as missing
    use_llm = False
    if args.no_llm:
        use_llm = False
    elif args.use_llm:
        if api_key:
            use_llm = True
        else:
            print("\n  ⚠  --use-llm specified but ANTHROPIC_API_KEY not set. Falling back to keywords.")
    elif api_key:
        use_llm = True  # Auto-detect: key present → use LLM

    if use_llm:
        print(f"\n  🧠 Classification: Claude LLM (Haiku)  |  Cache: {args.cache_path}")
    else:
        print(f"\n  📋 Classification: Keyword matching")
        if not api_key:
            print(f"     (Set ANTHROPIC_API_KEY env var to enable LLM classification)")

    # ── Viewer HTML (run once to create the permanent viewer) ──────────────────
    if args.viewer:
        viewer_path = Path(args.viewer)
        viewer_path.parent.mkdir(parents=True, exist_ok=True)
        viewer_html = generate_html([], generated, viewer_mode=True, github_mode=True)
        viewer_path.write_text(viewer_html, encoding="utf-8")
        size_kb = viewer_path.stat().st_size // 1024
        print(f"\n  ✓ Permanent viewer saved → {viewer_path}  ({size_kb} KB)")
        print(f"    This file never needs to be rebuilt.")
        print(f"    When hosted on GitHub Pages, data auto-loads from ./data/trials_data.json\n")
        if not args.json and not args.output:
            return

    # ── Data pipeline: fetch / parse ──────────────────────────────────────────
    since_date = until_date = None
    if not args.viewer or args.json or args.output:
        if args.csv:
            records = parse_csv_file(args.csv)
        else:
            max_pages = 1 if args.test else args.max_pages
            raw, since_date, until_date = fetch_all_studies(max_pages=max_pages, days=args.days)
            records = process_studies(raw)

        if not records:
            print("  ✗ No records to process. Exiting.")
            return

    # ── Classification ────────────────────────────────────────────────────────
    if use_llm:
        print(f"\n{'='*62}")
        print(f"  LLM Classification (Claude Haiku)")
        print(f"{'='*62}")
        try:
            records = classify_with_claude(records, args.cache_path, api_key)
        except Exception as e:
            print(f"  ✗ LLM classification failed: {e}")
            print(f"  Falling back to keyword classification…")
            for r in records:
                if not r.get("therapy_areas") or r["therapy_areas"] == []:
                    r["therapy_areas"] = classify_therapy_areas(
                        r.get("conditions", []), r.get("title", ""),
                        r.get("summary", ""), r.get("interventions", [])
                    )

    # ── JSON output (scheduled refresh mode) ──────────────────────────────────
    if args.json:
        json_path = Path(args.json)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_data = generate_json(records, generated, days=args.days,
                                   since_date=since_date, until_date=until_date)
        json_path.write_text(json_data, encoding="utf-8")
        size_kb = json_path.stat().st_size // 1024
        print(f"\n  ✓ Data saved → {json_path}  ({size_kb} KB, {len(records):,} trials)")
        print(f"    Load this file in the dashboard viewer to explore.\n")
        if not args.output or args.output == str(OUTPUT_HTML):
            return   # Don't also write the self-contained HTML unless explicitly asked

    # ── Self-contained HTML (legacy / one-off sharing mode) ───────────────────
    if not args.json or args.output != str(OUTPUT_HTML):
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        html = generate_html(records, generated, days=args.days,
                              since_date=since_date, until_date=until_date)
        out_path.write_text(html, encoding="utf-8")
        size_kb = out_path.stat().st_size // 1024
        print(f"\n  ✓ Self-contained dashboard → {out_path}  ({size_kb} KB, {len(records):,} trials)")
        print(f"    Open in browser: file://{out_path}\n")

if __name__ == "__main__":
    main()
